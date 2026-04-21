# app.py - Dashboard de Ventas Farmacéuticas

from dotenv import load_dotenv
import os

# Cargar variables de entorno ANTES de importar módulos que las necesitan
# override=True: Sobrescribir variables del sistema con valores del .env
load_dotenv(override=True)

# Configurar logging ANTES de importar otros módulos
from src.logging_config import setup_logging, get_logger
setup_logging(log_level=os.getenv('LOG_LEVEL', 'INFO'))
logger = get_logger(__name__)

from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, g, jsonify
from src.odoo_manager import OdooManager
from src.supabase_manager import SupabaseManager
from src.analytics_supabase import AnalyticsSupabase
from src.permissions_manager import PermissionsManager
from src.audit_logger import AuditLogger
from src.utils import get_meses_del_año, normalizar_linea_comercial, limpiar_nombre_producto, limpiar_nombre_atrevia
from authlib.integrations.flask_client import OAuth
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from functools import wraps
import pandas as pd
import json
import io
import calendar
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import pytz

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY')

# Configuración de Rate Limiting (A01: Broken Authentication - ISO/IEC 27001:2022 A.9.4.2)
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri=os.getenv('REDIS_URL', 'memory://'),  # Usa Redis si está disponible, sino memoria
    strategy='fixed-window'
)

# Configuración de seguridad para sesiones (A01: Broken Authentication)
app.config.update(
    # Expiración de sesión: 8 horas de inactividad
    PERMANENT_SESSION_LIFETIME=timedelta(hours=8),
    
    # Cookies de sesión no accesibles desde JavaScript (previene XSS)
    SESSION_COOKIE_HTTPONLY=True,
    
    # Protección CSRF: cookies solo enviadas con requests same-site
    SESSION_COOKIE_SAMESITE='Lax',  # 'Lax' permite OAuth redirects
    
    # SECURE solo en producción (requiere HTTPS)
    # En desarrollo con HTTP, debe estar en False
    SESSION_COOKIE_SECURE=os.getenv('FLASK_ENV') == 'production',
    
    # Nombre de cookie personalizado (security by obscurity)
    SESSION_COOKIE_NAME='__Host-session' if os.getenv('FLASK_ENV') == 'production' else 'session'
)

# Configuración OAuth2 Google
oauth = OAuth(app)
google = oauth.register(
    name='google',
    client_id=os.getenv('GOOGLE_CLIENT_ID'),
    client_secret=os.getenv('GOOGLE_CLIENT_SECRET'),
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    client_kwargs={
        'scope': 'openid email profile'
    }
)

# Timezone de Perú
PERU_TZ = pytz.timezone('America/Lima')
UTC_TZ = pytz.UTC

# Configuración para deshabilitar cache de templates
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

# --- Filtros Personalizados de Jinja2 ---

@app.template_filter('to_peru_time')
def to_peru_time(utc_timestamp_str):
    """
    Convierte timestamp UTC (ISO string) a hora de Perú (UTC-5).
    
    Args:
        utc_timestamp_str: String ISO 8601 en UTC (ej: "2026-04-21T15:57:42+00:00")
    
    Returns:
        str: Fecha y hora en formato "DD/MM/YYYY HH:MM:SS" (hora de Perú)
    """
    try:
        # Parse del timestamp UTC
        if isinstance(utc_timestamp_str, str):
            # Remover 'Z' o '+00:00' y parsear
            utc_timestamp_str = utc_timestamp_str.replace('Z', '+00:00')
            utc_dt = datetime.fromisoformat(utc_timestamp_str)
        else:
            utc_dt = utc_timestamp_str
        
        # Si no tiene timezone, asumir UTC
        if utc_dt.tzinfo is None:
            utc_dt = UTC_TZ.localize(utc_dt)
        
        # Convertir a hora de Perú
        peru_dt = utc_dt.astimezone(PERU_TZ)
        
        # Formatear como "21/04/2026 10:57:42"
        return peru_dt.strftime('%d/%m/%Y %H:%M:%S')
    except Exception as e:
        logger.error(f"Error convirtiendo timestamp a hora de Perú: {e}")
        return utc_timestamp_str  # Retornar original si falla

@app.template_filter('to_peru_date')
def to_peru_date(utc_timestamp_str):
    """
    Convierte timestamp UTC a solo fecha en hora de Perú.
    
    Returns:
        str: Fecha en formato "DD/MM/YYYY"
    """
    try:
        if isinstance(utc_timestamp_str, str):
            utc_timestamp_str = utc_timestamp_str.replace('Z', '+00:00')
            utc_dt = datetime.fromisoformat(utc_timestamp_str)
        else:
            utc_dt = utc_timestamp_str
        
        if utc_dt.tzinfo is None:
            utc_dt = UTC_TZ.localize(utc_dt)
        
        peru_dt = utc_dt.astimezone(PERU_TZ)
        return peru_dt.strftime('%d/%m/%Y')
    except Exception as e:
        logger.error(f"Error convirtiendo fecha: {e}")
        return utc_timestamp_str

@app.template_filter('to_peru_time_only')
def to_peru_time_only(utc_timestamp_str):
    """
    Convierte timestamp UTC a solo hora en hora de Perú.
    
    Returns:
        str: Hora en formato "HH:MM:SS"
    """
    try:
        if isinstance(utc_timestamp_str, str):
            utc_timestamp_str = utc_timestamp_str.replace('Z', '+00:00')
            utc_dt = datetime.fromisoformat(utc_timestamp_str)
        else:
            utc_dt = utc_timestamp_str
        
        if utc_dt.tzinfo is None:
            utc_dt = UTC_TZ.localize(utc_dt)
        
        peru_dt = utc_dt.astimezone(PERU_TZ)
        return peru_dt.strftime('%H:%M:%S')
    except Exception as e:
        logger.error(f"Error convirtiendo hora: {e}")
        return utc_timestamp_str

@app.template_filter('role_display')
def role_display(role_code):
    """
    Convierte código de rol a nombre legible en español.
    
    Args:
        role_code: Código del rol (admin_full, user_basic, etc)
    
    Returns:
        str: Nombre del rol en español
    """
    role_names = {
        'admin_full': 'Admin Full',
        'admin_export': 'Admin Export',
        'analytics_viewer': 'Analytics',
        'user_basic': 'Usuario Básico'
    }
    return role_names.get(role_code, role_code)

# --- Security Headers y CORS (A04: Insecure Design) ---

@app.after_request
def add_security_headers(response):
    """
    Añade headers de seguridad a todas las respuestas HTTP.
    - X-Frame-Options: Previene clickjacking
    - X-Content-Type-Options: Previene MIME sniffing
    - X-XSS-Protection: Protección XSS para navegadores legacy
    - Content-Security-Policy: Control de recursos cargados
    - Strict-Transport-Security: Fuerza HTTPS (solo en producción)
    - Referrer-Policy: Control de información en headers
    """
    
    # Previene que la página sea cargada en un iframe (clickjacking)
    # SAMEORIGIN permite iframes del mismo dominio (útil para OAuth redirects)
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    
    # Previene MIME type sniffing (fuerza a navegador a respetar Content-Type)
    response.headers['X-Content-Type-Options'] = 'nosniff'
    
    # Protección XSS para navegadores legacy (IE, Safari antiguo)
    response.headers['X-XSS-Protection'] = '1; mode=block'
    
    # Content Security Policy - Configuración permisiva para desarrollo
    # Permitimos inline scripts/styles y recursos de CDNs conocidos
    # TODO: Hacer más estricto en producción moviendo scripts inline a archivos externos
    csp_directives = [
        "default-src 'self'",  # Por defecto solo recursos del mismo origen
        "script-src 'self' 'unsafe-inline' 'unsafe-eval' https://accounts.google.com https://*.google.com https://www.gstatic.com https://cdn.jsdelivr.net https://cdnjs.cloudflare.com https://cdn.datatables.net https://code.jquery.com",  # Scripts + Google + Google Charts + CDNs + DataTables
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://www.gstatic.com https://cdn.jsdelivr.net https://cdnjs.cloudflare.com https://cdn.datatables.net",  # Estilos + Google Fonts + Google Charts + CDNs + DataTables
        "img-src 'self' data: https: blob:",  # Imágenes: locales, data URIs, HTTPS y blobs
        "font-src 'self' data: https://fonts.gstatic.com https://cdn.jsdelivr.net https://cdnjs.cloudflare.com",  # Fuentes de Google y CDNs
        "connect-src 'self' https://accounts.google.com https://*.google.com https://*.gstatic.com https://cdn.jsdelivr.net https://cdnjs.cloudflare.com https://cdn.datatables.net",  # Conexiones AJAX/fetch a Google y CDNs + DataTables
        "frame-src 'self' https://accounts.google.com https://*.google.com",  # iframes de Google OAuth
        "frame-ancestors 'self'",  # Solo el mismo origen puede embeber esta página
        "base-uri 'self'",  # Previene inyección de base tag
        "form-action 'self' https://accounts.google.com"  # Forms a mismo origen + Google OAuth
    ]
    response.headers['Content-Security-Policy'] = "; ".join(csp_directives)
    
    # Strict-Transport-Security: fuerza HTTPS (solo en producción)
    # max-age=31536000 = 1 año
    # includeSubDomains: aplica a todos los subdominios
    if os.getenv('FLASK_ENV') == 'production':
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    
    # Referrer-Policy: controla qué información se envía en el header Referer
    # strict-origin-when-cross-origin: envía origin completo en mismo origen,
    # solo origin en cross-origin, nada si de HTTPS a HTTP
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    
    # CORS: permitir solo mismo origen (por defecto)
    # Si necesitas permitir dominios específicos, configúralo en .env
    allowed_origins = os.getenv('CORS_ALLOWED_ORIGINS', '').split(',')
    origin = request.headers.get('Origin')
    
    if origin and origin in allowed_origins:
        response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Access-Control-Allow-Credentials'] = 'true'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    
    return response

# --- Inicialización de Managers ---
try:
    data_manager = OdooManager()
    logger.info(f"✅ data_manager inicializado: {type(data_manager).__name__}")
    logger.info(f"✅ OdooManager UID: {getattr(data_manager, 'uid', 'N/A')}")
except Exception as e:
    logger.warning(f"❌ No se pudo inicializar OdooManager: {e}. Continuando en modo offline.")
    # Crear un stub mínimo con las funciones usadas en la app para evitar fallos
    class _StubManager:
        def get_filter_options(self):
            return {'lineas': [], 'clients': []}
        def get_sales_lines(self, *args, **kwargs):
            logger.warning("⚠️ Usando _StubManager (modo offline) - retornando []")
            return []
        def get_all_sellers(self):
            return []
        def get_commercial_lines_stacked_data(self, *args, **kwargs):
            return {'yAxis': [], 'series': [], 'legend': []}
    data_manager = _StubManager()
    logger.warning(f"⚠️ data_manager en modo stub: {type(data_manager).__name__}")

# Inicializar Supabase Manager para metas de 2026
supabase_manager = SupabaseManager()

# Inicializar sistema de analytics (Supabase)
analytics_db = AnalyticsSupabase()

# Inicializar sistema de permisos
permissions_manager = PermissionsManager()

# Inicializar sistema de auditoría
audit_logger = AuditLogger()

# --- Context Processor para Templates ---
@app.context_processor
def inject_datetime():
    """Inyectar funciones de fecha/hora en todas las plantillas"""
    return {
        'now': datetime.now,
        'datetime': datetime
    }

# --- Middleware para Analytics y Seguridad ---

def verify_session_expiration():
    """
    Verifica que la sesión no haya expirado por:
    1. Timeout de INACTIVIDAD: 15 minutos sin actividad
    2. Timeout ABSOLUTO: 8 horas desde login (seguridad)
    """
    # Solo verificar si el usuario está logueado
    if 'username' not in session:
        return True
    
    # Inicializar timestamps si no existen (sesiones antiguas)
    if 'login_time' not in session:
        session['login_time'] = datetime.now(UTC_TZ).isoformat()
        session['last_activity_time'] = datetime.now(UTC_TZ).isoformat()
        return True
    
    try:
        now = datetime.now(UTC_TZ)
        username = session.get('username', 'unknown')
        
        # 1. Verificar TIMEOUT ABSOLUTO (desde login)
        login_time = datetime.fromisoformat(session['login_time'])
        if login_time.tzinfo is None:
            login_time = UTC_TZ.localize(login_time)
        
        elapsed_since_login = now - login_time
        max_session_hours = int(os.getenv('MAX_ABSOLUTE_SESSION_HOURS', '8'))
        
        if elapsed_since_login > timedelta(hours=max_session_hours):
            logger.warning(f"Sesión expirada por timeout absoluto para {username} (duración: {elapsed_since_login})")
            
            # Registrar timeout absoluto en auditoría
            audit_logger.log_session_timeout(
                user_email=username,
                timeout_type='absolute_limit',
                last_activity=session.get('last_activity_time'),
                ip_address=request.remote_addr if request else None
            )
            
            return False
        
        # 2. Verificar TIMEOUT DE INACTIVIDAD (desde última actividad)
        if 'last_activity_time' in session:
            last_activity = datetime.fromisoformat(session['last_activity_time'])
            if last_activity.tzinfo is None:
                last_activity = UTC_TZ.localize(last_activity)
            
            elapsed_since_activity = now - last_activity
            max_idle_minutes = int(os.getenv('MAX_IDLE_MINUTES', '15'))
            
            if elapsed_since_activity > timedelta(minutes=max_idle_minutes):
                logger.warning(f"Sesión expirada por inactividad para {username} (inactivo: {elapsed_since_activity})")
                
                # Registrar timeout por inactividad en auditoría
                audit_logger.log_session_timeout(
                    user_email=username,
                    timeout_type='inactivity',
                    last_activity=session.get('last_activity_time'),
                    ip_address=request.remote_addr if request else None
                )
                
                return False
        
        return True
    except Exception as e:
        logger.error(f"Error verificando expiración de sesión: {e}", exc_info=True)
        # En caso de error, permitir continuar (fail-open para desarrollo)
        return True

@app.before_request
def before_request():
    """Registra información de la petición antes de procesarla."""
    g.start_time = datetime.now()
    
    # Verificar expiración de sesión (opcional con env var)
    if os.getenv('ENABLE_SESSION_EXPIRATION', 'false').lower() == 'true':
        if 'username' in session:
            if not verify_session_expiration():
                session.clear()
                flash('Tu sesión ha expirado. Por favor, inicia sesión nuevamente.', 'warning')
                return redirect(url_for('login'))
            else:
                # Actualizar timestamp de última actividad en cada request válido
                session['last_activity_time'] = datetime.now(UTC_TZ).isoformat()
                session.modified = True  # Forzar que Flask guarde la sesión

@app.after_request
def after_request(response):
    """Registra la visita después de procesar la petición."""
    # Solo registrar si el usuario está logueado y la petición es exitosa
    if 'username' in session and response.status_code == 200:
        # Excluir cuenta de administrador de estadísticas (para deployments)
        excluded_users = ['jonathan.cerda@agrovetmarket.com']
        
        if session.get('username') not in excluded_users:
            # No registrar peticiones a archivos estáticos
            if not request.path.startswith('/static/'):
                try:
                    # Mapeo de rutas a títulos
                    page_titles = {
                        '/': 'Dashboard Principal',
                        '/dashboard': 'Dashboard de Ventas',
                        '/equipo-ventas': 'Equipo de Ventas',
                        '/meta': 'Metas de Ventas',
                        '/sales': 'Ventas Detalladas',
                        '/metas-vendedor': 'Metas por Vendedor',
                        '/dashboard-linea': 'Dashboard por Línea',
                        '/analytics': 'Analytics y Estadísticas'
                    }
                    
                    page_title = page_titles.get(request.path, request.path)
                    
                    analytics_db.log_visit(
                        user_email=session.get('username'),
                        user_name=session.get('user_name'),
                        page_url=request.path,
                        page_title=page_title,
                        ip_address=request.remote_addr,
                        user_agent=request.headers.get('User-Agent'),
                        referrer=request.referrer,
                        method=request.method
                    )
                except Exception as e:
                    logger.error(f"Error al registrar analytics: {e}", exc_info=True)
    
    return response

# --- Decoradores de Autenticación y Autorización ---

def login_required(f):
    """Decorador que requiere que el usuario esté logueado"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            flash('Por favor inicia sesión para acceder', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def require_admin_full(f):
    """Decorador que requiere rol admin_full para acceder"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            flash('Por favor inicia sesión para acceder', 'warning')
            return redirect(url_for('login'))
        
        user_email = session.get('username')
        if not permissions_manager.is_admin(user_email):
            flash('Acceso denegado. Esta sección requiere permisos de administrador.', 'danger')
            logger.warning(f"Intento de acceso no autorizado a ruta admin por {user_email}")
            return redirect(url_for('dashboard'))
        
        return f(*args, **kwargs)
    return decorated_function

# --- Rutas del Módulo de Administración de Permisos ---

@app.route('/admin/users')
@require_admin_full
def admin_users():
    """
    Lista todos los usuarios con sus roles y permisos.
    Solo accesible por usuarios con rol admin_full.
    """
    try:
        # Obtener todos los usuarios
        users = permissions_manager.get_all_users(include_inactive=False)
        
        # Obtener logs recientes para el dashboard
        recent_logs = audit_logger.get_recent_logs(limit=10)
        
        # Estadísticas
        total_users = permissions_manager.count_users(active_only=True)
        total_admins = permissions_manager.count_admins(active_only=True)
        changes_last_week = audit_logger.count_changes_last_week()
        
        # Obtener todos los roles disponibles para filtros
        roles = permissions_manager.get_all_roles()
        
        return render_template('admin/users_list.html',
                             users=users,
                             recent_logs=recent_logs,
                             total_users=total_users,
                             total_admins=total_admins,
                             changes_last_week=changes_last_week,
                             roles=roles,
                             current_user=session.get('username'))
    except Exception as e:
        logger.error(f"Error en ruta /admin/users: {e}", exc_info=True)
        flash('Error al cargar lista de usuarios', 'danger')
        return redirect(url_for('dashboard'))

@app.route('/admin/users/add', methods=['GET', 'POST'])
@require_admin_full
def admin_add_user():
    """
    Formulario para agregar nuevo usuario.
    GET: Muestra formulario
    POST: Procesa creación de usuario
    """
    if request.method == 'POST':
        try:
            email = request.form.get('email', '').strip().lower()
            role = request.form.get('role', '').strip()
            
            # Validaciones
            if not email or not role:
                flash('Email y rol son requeridos', 'danger')
                return redirect(url_for('admin_add_user'))
            
            # Validar email corporativo
            allowed_domains = os.getenv('ALLOWED_EMAIL_DOMAINS', '@agrovetmarket.com').split(',')
            if not any(email.endswith(domain) for domain in allowed_domains):
                flash(f'Email debe ser corporativo ({", ".join(allowed_domains)})', 'danger')
                return redirect(url_for('admin_add_user'))
            
            # Validar que el rol exista
            if role not in permissions_manager.ROLE_PERMISSIONS:
                flash('Rol no válido', 'danger')
                return redirect(url_for('admin_add_user'))
            
            # Verificar si el usuario ya existe
            existing_role = permissions_manager.get_user_role(email)
            if existing_role:
                flash(f'El usuario {email} ya existe con rol {existing_role}', 'warning')
                return redirect(url_for('admin_users'))
            
            # Crear usuario
            admin_email = session.get('username')
            success = permissions_manager.add_user(email, role, created_by=admin_email)
            
            if success:
                # Registrar en audit log
                audit_logger.log_user_created(
                    admin_email=admin_email,
                    new_user_email=email,
                    role=role,
                    ip_address=request.remote_addr,
                    user_agent=request.headers.get('User-Agent')
                )
                
                flash(f'Usuario {email} creado exitosamente con rol {role}', 'success')
                logger.info(f"Usuario {email} creado por {admin_email} con rol {role}")
            else:
                flash('Error al crear usuario', 'danger')
            
            return redirect(url_for('admin_users'))
            
        except Exception as e:
            logger.error(f"Error al crear usuario: {e}", exc_info=True)
            flash('Error al crear usuario', 'danger')
            return redirect(url_for('admin_add_user'))
    
    # GET: Mostrar formulario
    roles = permissions_manager.get_all_roles()
    return render_template('admin/user_add.html', roles=roles)

@app.route('/admin/users/edit/<email>', methods=['GET', 'POST'])
@require_admin_full
def admin_edit_user(email):
    """
    Formulario para editar rol de usuario existente.
    GET: Muestra formulario con datos actuales
    POST: Procesa actualización
    """
    # Prevenir que admin edite su propio rol
    current_user = session.get('username')
    if email.lower() == current_user.lower():
        flash('No puedes editar tu propio rol', 'warning')
        return redirect(url_for('admin_users'))
    
    if request.method == 'POST':
        try:
            new_role = request.form.get('role', '').strip()
            
            if not new_role:
                flash('Rol es requerido', 'danger')
                return redirect(url_for('admin_edit_user', email=email))
            
            # Validar que el rol exista
            if new_role not in permissions_manager.ROLE_PERMISSIONS:
                flash('Rol no válido', 'danger')
                return redirect(url_for('admin_edit_user', email=email))
            
            # Obtener rol actual
            old_role = permissions_manager.get_user_role(email)
            if not old_role:
                flash(f'Usuario {email} no encontrado', 'danger')
                return redirect(url_for('admin_users'))
            
            # No actualizar si es el mismo rol
            if old_role == new_role:
                flash('El usuario ya tiene ese rol', 'info')
                return redirect(url_for('admin_users'))
            
            # Actualizar rol
            success = permissions_manager.update_user_role(email, new_role)
            
            if success:
                # Registrar en audit log
                audit_logger.log_user_updated(
                    admin_email=current_user,
                    user_email=email,
                    old_role=old_role,
                    new_role=new_role,
                    ip_address=request.remote_addr,
                    user_agent=request.headers.get('User-Agent')
                )
                
                flash(f'Rol de {email} actualizado: {old_role} → {new_role}', 'success')
                logger.info(f"Rol de {email} actualizado por {current_user}: {old_role} → {new_role}")
            else:
                flash('Error al actualizar rol', 'danger')
            
            return redirect(url_for('admin_users'))
            
        except Exception as e:
            logger.error(f"Error al actualizar usuario: {e}", exc_info=True)
            flash('Error al actualizar usuario', 'danger')
            return redirect(url_for('admin_users'))
    
    # GET: Mostrar formulario
    user_details = permissions_manager.get_user_details(email)
    if not user_details:
        flash(f'Usuario {email} no encontrado', 'danger')
        return redirect(url_for('admin_users'))
    
    roles = permissions_manager.get_all_roles()
    return render_template('admin/user_edit.html', user=user_details, roles=roles)

@app.route('/admin/users/delete/<email>', methods=['POST'])
@require_admin_full
def admin_delete_user(email):
    """
    Elimina (desactiva) un usuario.
    Solo POST para prevenir eliminación accidental.
    """
    # Prevenir que admin se auto-elimine
    current_user = session.get('username')
    if email.lower() == current_user.lower():
        return jsonify({'success': False, 'message': 'No puedes eliminar tu propia cuenta'}), 403
    
    try:
        # Verificar que el usuario existe
        user_role = permissions_manager.get_user_role(email)
        if not user_role:
            return jsonify({'success': False, 'message': f'Usuario {email} no encontrado'}), 404
        
        # Eliminar usuario (soft delete por defecto)
        soft_delete = request.form.get('soft_delete', 'true').lower() == 'true'
        success = permissions_manager.delete_user(email, soft_delete=soft_delete)
        
        if success:
            # Registrar en audit log
            audit_logger.log_user_deleted(
                admin_email=current_user,
                user_email=email,
                soft_delete=soft_delete,
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent')
            )
            
            action = 'desactivado' if soft_delete else 'eliminado'
            flash(f'Usuario {email} {action} exitosamente', 'success')
            logger.info(f"Usuario {email} {action} por {current_user}")
            return jsonify({'success': True, 'message': f'Usuario {action} exitosamente'})
        else:
            return jsonify({'success': False, 'message': 'Error al eliminar usuario'}), 500
            
    except Exception as e:
        logger.error(f"Error al eliminar usuario: {e}", exc_info=True)
        return jsonify({'success': False, 'message': 'Error al eliminar usuario'}), 500

@app.route('/admin/users/reactivate/<email>', methods=['POST'])
@require_admin_full
def admin_reactivate_user(email):
    """Reactiva un usuario previamente desactivado"""
    current_user = session.get('username')
    
    try:
        success = permissions_manager.reactivate_user(email)
        
        if success:
            # Registrar en audit log
            audit_logger.log_user_reactivated(
                admin_email=current_user,
                user_email=email,
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent')
            )
            
            flash(f'Usuario {email} reactivado exitosamente', 'success')
            logger.info(f"Usuario {email} reactivado por {current_user}")
            return jsonify({'success': True, 'message': 'Usuario reactivado exitosamente'})
        else:
            return jsonify({'success': False, 'message': 'Error al reactivar usuario'}), 500
            
    except Exception as e:
        logger.error(f"Error al reactivar usuario: {e}", exc_info=True)
        return jsonify({'success': False, 'message': 'Error al reactivar usuario'}), 500

@app.route('/admin/audit-log')
@require_admin_full
def admin_audit_log():
    """
    Dashboard de auditoría con historial completo de cambios y métricas de seguridad.
    Solo accesible por admin_full.
    """
    try:
        # Parámetros de filtrado
        days = request.args.get('days', 30, type=int)
        action = request.args.get('action', '')
        admin_email = request.args.get('admin', '')
        
        logger.info(f"[AUDIT-LOG] Cargando dashboard de auditoría - days={days}, action={action}, admin={admin_email}")
        
        # Obtener logs de PERMISOS solamente (excluye eventos de autenticación)
        logger.info("[AUDIT-LOG] Obteniendo logs de permisos...")
        permissions_logs = audit_logger.get_filtered_logs(
            days=days, 
            action=action, 
            admin_email=admin_email,
            exclude_auth_events=True  # 🆕 Solo permisos (CREATE, UPDATE, DELETE, etc)
        )
        logger.info(f"[AUDIT-LOG] ✓ Logs de permisos obtenidos: {len(permissions_logs)} registros")
        
        # Estadísticas generales de permisos
        logger.info("[AUDIT-LOG] Obteniendo estadísticas generales...")
        stats = audit_logger.get_statistics()
        logger.info(f"[AUDIT-LOG] ✓ Stats obtenidas")
        
        # 🆕 Estadísticas de seguridad (login/logout) - últimas 24 horas
        logger.info("[AUDIT-LOG] Obteniendo estadísticas de seguridad...")
        security_stats = audit_logger.get_security_stats(hours=24)
        logger.info(f"[AUDIT-LOG] ✓ Security stats: {security_stats}")
        
        # 🆕 Timeline de login/logout para gráfica - últimas 24 horas
        logger.info("[AUDIT-LOG] Obteniendo timeline de login...")
        login_timeline = audit_logger.get_login_timeline(hours=24)
        logger.info(f"[AUDIT-LOG] ✓ Login timeline: {len(login_timeline)} horas")
        
        # 🆕 Intentos fallidos recientes (últimos 10)
        logger.info("[AUDIT-LOG] Obteniendo intentos fallidos...")
        failed_attempts = audit_logger.get_recent_failed_attempts(limit=10)
        logger.info(f"[AUDIT-LOG] ✓ Failed attempts: {len(failed_attempts)} registros")
        
        logger.info("[AUDIT-LOG] Renderizando template audit_log.html...")
        return render_template('admin/audit_log.html',
                             logs=permissions_logs,  # 🆕 Cambiado: ahora solo permisos
                             stats=stats,
                             security_stats=security_stats,
                             login_timeline=login_timeline,
                             failed_attempts=failed_attempts,
                             days=days,
                             current_action=action,
                             current_admin=admin_email)
    except Exception as e:
        logger.error(f"❌ Error en ruta /admin/audit-log: {e}", exc_info=True)
        logger.error(f"❌ Tipo de error: {type(e).__name__}")
        logger.error(f"❌ Args del error: {e.args}")
        flash(f'Error al cargar historial de auditoría: {str(e)}', 'danger')
        return redirect(url_for('admin_users'))

# --- Funciones Auxiliares ---

# Las funciones auxiliares han sido movidas a src/utils/ para mejor organización
# Imports: get_meses_del_año, normalizar_linea_comercial, limpiar_nombre_producto, limpiar_nombre_atrevia

@app.route('/login')
@limiter.limit("10 per minute")  # Máximo 10 accesos por minuto a la página de login
def login():
    """Mostrar página de login con botón de Google"""
    return render_template('login.html')

@app.route('/google-oauth')
@limiter.limit("10 per minute")  # Máximo 10 intentos de OAuth por minuto
def google_oauth():
    """Iniciar el proceso de autenticación con Google OAuth2"""
    redirect_uri = url_for('authorize', _external=True)
    return google.authorize_redirect(redirect_uri)

@app.route('/loading')
def loading():
    """Mostrar página de carga mientras se cargan los datos del dashboard"""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    # Si el usuario es Administrador Total, redirigir a Administración de Usuarios
    user_role = session.get('user_role')
    if user_role == 'admin_full':
        return redirect(url_for('admin_users'))
    
    return render_template('loading.html')

@app.route('/authorize')
@limiter.limit("5 per minute")  # Más restrictivo para OAuth callback (máximo 5 por minuto)
def authorize():
    """Callback de Google OAuth2 - Procesar la respuesta de autenticación"""
    try:
        # Obtener el token de acceso de Google
        token = google.authorize_access_token()
        
        # Obtener información del usuario desde Google
        user_info = token.get('userinfo')
        
        if user_info:
            email = user_info.get('email')
            name = user_info.get('name')
            
            # Verificar si el usuario tiene permisos en Supabase
            try:
                role = permissions_manager.get_user_role(email)
                
                if role:
                    # Usuario autenticado y autorizado (tiene rol en Supabase)
                    session.permanent = True  # Habilitar expiración con PERMANENT_SESSION_LIFETIME
                    session['username'] = email
                    session['user_name'] = name
                    session['user_info'] = user_info
                    session['user_role'] = role  # Guardar rol en sesión
                    session['login_time'] = datetime.now(UTC_TZ).isoformat()  # Timestamp de login
                    session['last_activity_time'] = datetime.now(UTC_TZ).isoformat()  # Timestamp de última actividad
                    session['session_id'] = os.urandom(16).hex()  # ID único de sesión
                    
                    # Registrar login exitoso en auditoría
                    audit_logger.log_login_success(
                        user_email=email,
                        user_name=name,
                        role=role,
                        ip_address=request.remote_addr,
                        user_agent=request.headers.get('User-Agent'),
                        oauth_provider='google',
                        session_id=session.get('session_id')
                    )
                    
                    logger.info(f"Usuario autenticado: {email} (rol: {role})")
                    flash('¡Inicio de sesión exitoso!', 'success')
                    return redirect(url_for('loading'))
                else:
                    # Usuario autenticado pero no autorizado
                    logger.warning(f"Intento de acceso denegado: {email} (sin rol en sistema)")
                    
                    # Registrar intento fallido de login en auditoría
                    audit_logger.log_login_failed(
                        attempted_email=email,
                        ip_address=request.remote_addr,
                        user_agent=request.headers.get('User-Agent'),
                        failure_reason='user_not_authorized',
                        error_message=f'El correo {email} no tiene permiso para acceder a esta aplicación.'
                    )
                    
                    flash(f'El correo {email} no tiene permiso para acceder a esta aplicación.', 'warning')
                    return redirect(url_for('login'))
                    
            except Exception as e:
                logger.error(f"Error verificando permisos para {email}: {e}", exc_info=True)
                
                # Registrar error en verificación de permisos
                audit_logger.log_login_failed(
                    attempted_email=email,
                    ip_address=request.remote_addr,
                    user_agent=request.headers.get('User-Agent'),
                    failure_reason='permission_check_error',
                    error_message=str(e)
                )
                
                flash('Error al verificar permisos. Por favor, contacte al administrador.', 'danger')
                return redirect(url_for('login'))
        else:
            # No se pudo obtener información del usuario desde Google
            logger.warning(f"OAuth error: No se pudo obtener información del usuario")
            
            audit_logger.log_login_failed(
                attempted_email=None,
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent'),
                failure_reason='oauth_userinfo_error',
                error_message='No se pudo obtener información del usuario de Google'
            )
            
            flash('No se pudo obtener información del usuario de Google.', 'danger')
            return redirect(url_for('login'))
            
    except Exception as e:
        logger.error(f"Error en autenticación OAuth: {e}", exc_info=True)
        
        # Registrar error general de OAuth
        audit_logger.log_login_failed(
            attempted_email=None,
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent'),
            failure_reason='oauth_error',
            error_message=str(e)
        )
        
        flash('Error en la autenticación. Por favor, intente nuevamente.', 'danger')
        return redirect(url_for('login'))

@app.route('/desing-login')
def desing_login():
    return render_template('desing_login.html')


@app.route('/logout')
def logout():
    """Cerrar sesión del usuario con logging de seguridad"""
    username = session.get('username', 'unknown')
    login_time_str = session.get('login_time')
    session_id = session.get('session_id')
    
    # Calcular duración de la sesión
    session_duration = None
    if login_time_str:
        try:
            login_time = datetime.fromisoformat(login_time_str)
            if login_time.tzinfo is None:
                login_time = UTC_TZ.localize(login_time)
            
            logout_time = datetime.now(UTC_TZ)
            duration_delta = logout_time - login_time
            
            # Formatear duración como HH:MM:SS
            hours, remainder = divmod(duration_delta.total_seconds(), 3600)
            minutes, seconds = divmod(remainder, 60)
            session_duration = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"
        except Exception as e:
            logger.warning(f"Error calculando duración de sesión: {e}")
    
    # Registrar logout en auditoría
    if username != 'unknown':
        audit_logger.log_logout(
            user_email=username,
            ip_address=request.remote_addr,
            session_duration=session_duration,
            logout_type='manual',
            session_id=session_id
        )
    
    logger.info(f"Logout: {username} cerró sesión desde {request.remote_addr} (duración: {session_duration or 'N/A'})")
    
    # Limpiar sesión
    session.clear()
    flash('Has cerrado sesión correctamente.', 'info')
    return redirect(url_for('login'))

@app.route('/')
def index():
    # Redirigir la ruta raíz al dashboard
    return redirect(url_for('dashboard'))

@app.route('/sales', methods=['GET', 'POST'])
def sales():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    # --- Verificación de Permisos ---
    username = session.get('username')
    if not permissions_manager.has_permission(username, 'view_analytics'):
        flash('No tienes permiso para acceder a esta página.', 'warning')
        return redirect(url_for('dashboard'))
    is_admin = permissions_manager.is_admin(username)
    # --- Fin Verificación ---
    
    try:
        # Obtener opciones de filtro
        filter_options = data_manager.get_filter_options()
        
        if request.method == 'POST':
            # For POST, get filters from the form
            selected_filters = {
                'date_from': request.form.get('date_from') or None,
                'date_to': request.form.get('date_to') or None,
                'search_term': request.form.get('search_term') or None
            }
        else:
            # For GET, start with no filters, so defaults will be used
            selected_filters = {
                'date_from': request.args.get('date_from') or None,
                'date_to': request.args.get('date_to') or None,
                'search_term': request.args.get('search_term') or None
            }

        # Create a clean copy for the database query
        query_filters = selected_filters.copy()

        # Clean up filter values for the query
        for key, value in query_filters.items():
            if not value:  # Handles empty strings and None
                query_filters[key] = None
        
        # Fetch data on every page load (GET and POST)
        # On GET, filters are None, so odoo_manager will use defaults (last 30 days)
        logger.info(f"🔍 Obteniendo líneas de venta completas...")
        logger.info(f"   Filtros: date_from={query_filters.get('date_from')}, date_to={query_filters.get('date_to')}")
        sales_data = data_manager.get_sales_lines(
            date_from=query_filters.get('date_from'),
            date_to=query_filters.get('date_to'),
            partner_id=None,
            search=query_filters.get('search_term'),
            linea_id=None,
            limit=1000
        )
        logger.info(f"📊 Obtenidas {len(sales_data)} líneas de ventas desde Odoo")
        
        # Filtrar VENTA INTERNACIONAL (exportaciones)
        sales_data_filtered = []
        for sale in sales_data:
            linea_comercial = sale.get('commercial_line_national_id')
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                nombre_linea = linea_comercial[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_linea:
                    continue
            
            # También filtrar por canal de ventas
            canal_ventas = sale.get('sales_channel_id')
            if canal_ventas and isinstance(canal_ventas, list) and len(canal_ventas) > 1:
                nombre_canal = canal_ventas[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_canal or 'INTERNACIONAL' in nombre_canal:
                    continue
            
            sales_data_filtered.append(sale)
        
        return render_template('sales.html', 
                             sales_data=sales_data_filtered,
                             filter_options=filter_options,
                             selected_filters=selected_filters,
                             fecha_actual=datetime.now(),
                             is_admin=is_admin) # Pasar el flag a la plantilla
    
    except Exception as e:
        logger.error(f"Error al obtener datos de ventas: {e}", exc_info=True)
        flash('No se pudieron cargar los datos de ventas. Intente nuevamente más tarde.', 'danger')
        return render_template('sales.html', 
                             sales_data=[],
                             filter_options={'lineas': [], 'clientes': []},
                             selected_filters={},
                             fecha_actual=datetime.now(),
                             is_admin=is_admin) # Pasar el flag a la plantilla

@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    try:
        # --- Lógica de Permisos de Administrador ---
        username = session.get('username')
        is_admin = permissions_manager.is_admin(username)

        # Obtener año seleccionado (parámetro o año actual por defecto)
        fecha_actual = datetime.now()
        año_seleccionado = request.args.get('año', str(fecha_actual.year))
        try:
            año_seleccionado = int(año_seleccionado)
        except (ValueError, TypeError):
            año_seleccionado = fecha_actual.year
        
        # Generar lista de años disponibles (desde 2025 hasta año actual)
        años_disponibles = list(range(2025, fecha_actual.year + 1))
        
        # Obtener mes seleccionado
        mes_seleccionado = request.args.get('mes', f"{año_seleccionado}-{fecha_actual.month:02d}" if año_seleccionado == fecha_actual.year else f"{año_seleccionado}-01")
        
        # --- NUEVA LÓGICA DE FILTRADO POR DÍA ---
        # Obtener el día final del filtro, si existe
        dia_fin_param = request.args.get('dia_fin')

        # Crear todos los meses del año seleccionado
        meses_disponibles = get_meses_del_año(año_seleccionado)
        
        # Obtener nombre del mes seleccionado
        mes_obj = next((m for m in meses_disponibles if m['key'] == mes_seleccionado), None)
        mes_nombre = mes_obj['nombre'] if mes_obj else "Mes Desconocido"
        
        año_sel, mes_sel = mes_seleccionado.split('-')
        
        # Determinar el día a usar para los cálculos y la fecha final
        if dia_fin_param:
            try:
                dia_actual = int(dia_fin_param)
                fecha_fin = f"{año_sel}-{mes_sel}-{str(dia_actual).zfill(2)}"
            except (ValueError, TypeError):
                # Si el parámetro no es un número válido, usar el comportamiento por defecto
                dia_fin_param = None # Resetear para que entre al siguiente bloque
        
        if not dia_fin_param:
            # Comportamiento original si no hay filtro de día
            if mes_seleccionado == fecha_actual.strftime('%Y-%m'):
                # Mes actual: usar día actual
                dia_actual = fecha_actual.day
            else:
                # Mes pasado: usar último día del mes
                ultimo_dia_mes = calendar.monthrange(int(año_sel), int(mes_sel))[1]
                dia_actual = ultimo_dia_mes
            fecha_fin = f"{año_sel}-{mes_sel}-{str(dia_actual).zfill(2)}"

        fecha_inicio = f"{año_sel}-{mes_sel}-01"
        # --- FIN DE LA NUEVA LÓGICA ---

        # Obtener metas del mes seleccionado desde la sesión
        metas_historicas = supabase_manager.read_metas_por_linea()
        metas_del_mes_raw = metas_historicas.get(mes_seleccionado, {}).get('metas', {})
        metas_ipn_del_mes_raw = metas_historicas.get(mes_seleccionado, {}).get('metas_ipn', {})
        
        # Normalizar claves a minúsculas para coincidir con los IDs de líneas
        metas_del_mes_raw = {k.lower(): v for k, v in metas_del_mes_raw.items()}
        metas_ipn_del_mes_raw = {k.lower(): v for k, v in metas_ipn_del_mes_raw.items()}
        
        # Consolidar metas de GENVET con TERCEROS
        metas_del_mes = {}
        metas_ipn_del_mes = {}
        
        for linea_id, valor in metas_del_mes_raw.items():
            if linea_id == 'genvet':
                # Sumar la meta de genvet a terceros
                metas_del_mes['terceros'] = metas_del_mes.get('terceros', 0) + valor
            else:
                metas_del_mes[linea_id] = valor
        
        for linea_id, valor in metas_ipn_del_mes_raw.items():
            if linea_id == 'genvet':
                # Sumar la meta IPN de genvet a terceros
                metas_ipn_del_mes['terceros'] = metas_ipn_del_mes.get('terceros', 0) + valor
            else:
                metas_ipn_del_mes[linea_id] = valor
        
        # Las líneas comerciales se generan dinámicamente más adelante.
        
        # Obtener datos reales de ventas desde Odoo
        try:
            # Las fechas de inicio y fin ahora se calculan más arriba
            
            # Obtener datos de ventas reales desde Odoo
            sales_data = data_manager.get_sales_lines(
                date_from=fecha_inicio,
                date_to=fecha_fin,
                limit=10000
            )
            
            logger.info(f"Obtenidas {len(sales_data)} líneas de ventas para el dashboard")
            
        except Exception as e:
            logger.error(f"Error obteniendo datos de Odoo: {e}", exc_info=True)
            sales_data = []
        
        # Procesar datos de ventas por línea comercial
        datos_lineas = []
        total_venta = 0
        total_vencimiento = 0
        total_venta_pn = 0
        
        # --- CÁLCULO DE TOTALES ---
        # Calcular totales de metas ANTES de filtrar las líneas para la tabla.
        # Esto asegura que ECOMMERCE se incluya en el total general del KPI.
        total_meta = sum(metas_del_mes.values())
        total_meta_pn = sum(metas_ipn_del_mes.values())
        
        # Mapeo de líneas comerciales de Odoo a IDs locales
        mapeo_lineas = {
            'PETMEDICA': 'petmedica',
            'AGROVET': 'agrovet', 
            'PET NUTRISCIENCE': 'pet_nutriscience',
            'AVIVET': 'avivet',
            'OTROS': 'otros',
            'TERCEROS': 'terceros',
            'INTERPET': 'interpet',
        }
        
        # Calcular ventas reales por línea comercial
        ventas_por_linea = {}
        ventas_por_ruta = {}
        ventas_ipn_por_linea = {} # Nueva variable para ventas de productos nuevos
        ventas_por_producto = {}
        ciclo_vida_por_producto = {}
        ventas_por_ciclo_vida = {}
        ventas_por_forma = {}
        for sale in sales_data:
            # Excluir VENTA INTERNACIONAL (exportaciones)
            linea_comercial = sale.get('commercial_line_national_id')
            nombre_linea_actual = None
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                nombre_linea_original = linea_comercial[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_linea_original:
                    continue
                # Aplicar normalización para agrupar GENVET y MARCA BLANCA como TERCEROS
                nombre_linea_actual = normalizar_linea_comercial(nombre_linea_original)
            
            # También filtrar por canal de ventas
            canal_ventas = sale.get('sales_channel_id')
            if canal_ventas and isinstance(canal_ventas, list) and len(canal_ventas) > 1:
                nombre_canal = canal_ventas[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_canal or 'INTERNACIONAL' in nombre_canal:
                    continue
            
            # Procesar el balance de la venta
            balance_float = float(sale.get('balance', 0))
            if balance_float != 0:
                
                # Sumar a ventas totales por línea
                if nombre_linea_actual:
                    ventas_por_linea[nombre_linea_actual] = ventas_por_linea.get(nombre_linea_actual, 0) + balance_float
                
                # LÓGICA FINAL: Sumar si la RUTA (route_id) coincide con los valores especificados
                ruta = sale.get('route_id')
                # Se cambia la comparación al ID de la ruta (ruta[0]) para evitar problemas con traducciones.
                if isinstance(ruta, list) and len(ruta) > 0 and ruta[0] in [18, 19]:
                    if nombre_linea_actual:
                        ventas_por_ruta[nombre_linea_actual] = ventas_por_ruta.get(nombre_linea_actual, 0) + balance_float
                
                # Sumar a ventas de productos nuevos (IPN) - Lógica restaurada
                ciclo_vida = sale.get('product_life_cycle')
                if ciclo_vida and ciclo_vida == 'nuevo':
                    if nombre_linea_actual:
                        ventas_ipn_por_linea[nombre_linea_actual] = ventas_ipn_por_linea.get(nombre_linea_actual, 0) + balance_float
                
                # Agrupar por producto para Top 7 (excluir GENVET, MARCA BLANCA y TERCEROS)
                producto_nombre = sale.get('name', '').strip()
                if producto_nombre and nombre_linea_actual not in ['GENVET', 'MARCA BLANCA', 'TERCEROS']:
                    # Limpiar nombres de ATREVIA eliminando indicadores de tamaño/presentación
                    producto_nombre_limpio = limpiar_nombre_atrevia(producto_nombre)
                    ventas_por_producto[producto_nombre_limpio] = ventas_por_producto.get(producto_nombre_limpio, 0) + balance_float
                    if producto_nombre_limpio not in ciclo_vida_por_producto:
                        ciclo_vida_por_producto[producto_nombre_limpio] = ciclo_vida
                
                # Agrupar por ciclo de vida para el gráfico de dona
                ciclo_vida_grafico = ciclo_vida if ciclo_vida else 'No definido'
                ventas_por_ciclo_vida[ciclo_vida_grafico] = ventas_por_ciclo_vida.get(ciclo_vida_grafico, 0) + balance_float

        print(f"💰 Ventas por línea comercial: {ventas_por_linea}")
        print(f"📦 Ventas por Vencimiento (Ciclo de Vida): {ventas_por_ruta}")
        print(f"✨ Ventas IPN (Productos Nuevos): {ventas_ipn_por_linea}")

        # --- Procesamiento de datos para gráficos (después del bucle) ---

        # 1. Procesar datos para la tabla principal
        # Generar dinámicamente las líneas comerciales a partir de ventas y metas
        all_lines = {}  # Usar un dict para evitar duplicados, con el id como clave

        # Añadir líneas desde las ventas reales
        for nombre_linea_venta in ventas_por_linea.keys():
            linea_id = nombre_linea_venta.lower().replace(' ', '_')
            all_lines[linea_id] = {'nombre': nombre_linea_venta.upper(), 'id': linea_id}

        # Añadir líneas desde las metas (para aquellas que no tuvieron ventas)
        # IMPORTANTE: Normalizar las claves de metas_del_mes a minúsculas para evitar duplicados
        for linea_id_meta_raw in metas_del_mes.keys():
            # Normalizar a minúsculas
            linea_id_meta = linea_id_meta_raw.lower()
            
            # Convertir genvet a terceros si existe en las metas
            if linea_id_meta == 'genvet':
                linea_id_meta = 'terceros'
            
            if linea_id_meta not in all_lines:
                # Reconstruir el nombre desde el ID de la meta
                nombre_reconstruido = linea_id_meta.replace('_', ' ').upper()
                all_lines[linea_id_meta] = {'nombre': nombre_reconstruido, 'id': linea_id_meta}
        
        # Convertir el diccionario de líneas a una lista ordenada por nombre
        lineas_comerciales_dinamicas = sorted(all_lines.values(), key=lambda x: x['nombre'])

        # Excluir líneas no deseadas que pueden venir de los datos
        lineas_a_excluir = ['LICITACION', 'NINGUNO', 'ECOMMERCE', 'GENVET', 'MARCA BLANCA']
        lineas_comerciales_filtradas = [
            linea for linea in lineas_comerciales_dinamicas
            if linea['nombre'].upper() not in lineas_a_excluir
        ]

        # Pre-calcular la venta total para el cálculo de porcentajes
        total_venta = sum(ventas_por_linea.values())
        total_venta_calculado = total_venta # Renombrar para claridad en el bucle
        
        print(f"🔍 DEBUG: lineas_comerciales_filtradas = {[l['nombre'] for l in lineas_comerciales_filtradas]}")
        print(f"🔍 DEBUG: total_venta = {total_venta}")

        for linea in lineas_comerciales_filtradas:
            meta = metas_del_mes.get(linea['id'], 0)
            nombre_linea = linea['nombre'].upper()
            
            # Usar ventas reales de Odoo
            venta = ventas_por_linea.get(nombre_linea, 0)
            print(f"🔍 DEBUG BUCLE: {nombre_linea} - meta={meta}, venta={venta}")
            
            # Usar la meta IPN registrada por el usuario
            meta_pn = metas_ipn_del_mes.get(linea['id'], 0)
            venta_pn = ventas_ipn_por_linea.get(nombre_linea, 0) # Usar el cálculo real de ventas de productos nuevos
            vencimiento = ventas_por_ruta.get(nombre_linea, 0) # Usamos el nuevo cálculo
            
            porcentaje_total = (venta / meta * 100) if meta > 0 else 0
            porcentaje_pn = (venta_pn / meta_pn * 100) if meta_pn > 0 else 0
            porcentaje_sobre_total = (venta / total_venta_calculado * 100) if total_venta_calculado > 0 else 0

            datos_lineas.append({
                'nombre': linea['nombre'],
                'meta': meta,
                'venta': venta, # Ahora es positivo
                'porcentaje_total': porcentaje_total,
                'porcentaje_sobre_total': porcentaje_sobre_total,
                'meta_pn': meta_pn,
                'venta_pn': venta_pn,
                'porcentaje_pn': porcentaje_pn,
                'vencimiento_6_meses': vencimiento
            })
            
            # Los totales de metas ya se calcularon. Aquí solo sumamos los totales de ventas.
            total_venta_pn += venta_pn
            total_vencimiento += vencimiento
        
        # --- 2. Calcular KPIs ---
        # Días laborables restantes (Lunes a Sábado)
        dias_restantes = 0
        ritmo_diario_requerido = 0
        if mes_seleccionado == fecha_actual.strftime('%Y-%m'):
            hoy = fecha_actual.day
            ultimo_dia_mes = calendar.monthrange(año_seleccionado, fecha_actual.month)[1]
            for dia in range(hoy, ultimo_dia_mes + 1):
                # weekday() -> Lunes=0, Domingo=6
                if datetime(año_seleccionado, fecha_actual.month, dia).weekday() < 6:
                    dias_restantes += 1
            
            porcentaje_restante = 100 - ((total_venta / total_meta * 100) if total_meta > 0 else 100)
            if porcentaje_restante > 0 and dias_restantes > 0:
                ritmo_diario_requerido = porcentaje_restante / dias_restantes

        # Calcular KPIs
        kpis = {
            'meta_total': total_meta,
            'venta_total': total_venta,
            'porcentaje_avance': (total_venta / total_meta * 100) if total_meta > 0 else 0,
            'meta_ipn': total_meta_pn,
            'venta_ipn': total_venta_pn,
            'porcentaje_avance_ipn': (total_venta_pn / total_meta_pn * 100) if total_meta_pn > 0 else 0,
            'vencimiento_6_meses': total_vencimiento,
            'avance_diario_total': ((total_venta / total_meta * 100) / dia_actual) if total_meta > 0 and dia_actual > 0 else 0,
            'avance_diario_ipn': ((total_venta_pn / total_meta_pn * 100) / dia_actual) if total_meta_pn > 0 and dia_actual > 0 else 0,
            'ritmo_diario_requerido': ritmo_diario_requerido
        }

        # --- Avance lineal: proyección de cierre y faltante ---
        # Proyección mensual lineal: proyectar ventas actuales al mes completo
        try:
            dias_en_mes = calendar.monthrange(int(año_sel), int(mes_sel))[1]
        except Exception:
            dias_en_mes = 30

        if dia_actual > 0:
            proyeccion_mensual = (total_venta / dia_actual) * dias_en_mes
        else:
            proyeccion_mensual = 0

        avance_lineal_pct = (proyeccion_mensual / total_meta * 100) if total_meta > 0 else 0
        faltante_meta = max(total_meta - total_venta, 0)

        # Cálculos específicos para IPN (usando las variables ya calculadas)
        # total_meta_pn ya está calculado arriba
        # total_venta_pn ya está calculado arriba
        
        # Proyección lineal IPN
        if dia_actual > 0:
            promedio_diario_ipn = total_venta_pn / dia_actual
            proyeccion_mensual_ipn = promedio_diario_ipn * dias_en_mes
        else:
            proyeccion_mensual_ipn = 0

        avance_lineal_ipn_pct = (proyeccion_mensual_ipn / total_meta_pn * 100) if total_meta_pn > 0 else 0
        faltante_meta_ipn = max(total_meta_pn - total_venta_pn, 0)

        
        # 3. Ordenar productos para el gráfico Top 7
        # Ordenar productos por ventas y tomar los top 7
        productos_ordenados = sorted(ventas_por_producto.items(), key=lambda x: x[1], reverse=True)[:7]
        
        datos_productos = []
        for nombre_producto, venta in productos_ordenados:
            datos_productos.append({
                'nombre': nombre_producto,
                'venta': venta,
                'ciclo_vida': ciclo_vida_por_producto.get(nombre_producto, 'No definido')
            })
        
        print(f"🏆 Top 7 productos por ventas: {[p['nombre'] for p in datos_productos]}")
        
        # 4. Ordenar datos para el gráfico de Ciclo de Vida
        # Convertir a lista ordenada por ventas
        datos_ciclo_vida = []
        for ciclo, venta in sorted(ventas_por_ciclo_vida.items(), key=lambda x: x[1], reverse=True):
            datos_ciclo_vida.append({
                'ciclo': ciclo,
                'venta': venta
            })
        
        print(f"📈 Ventas por Ciclo de Vida: {datos_ciclo_vida}")
        
        # --- INICIO: LÓGICA PARA LA TABLA DEL EQUIPO ECOMMERCE ---
        datos_ecommerce = []
        kpis_ecommerce = {'meta_total': 0, 'venta_total': 0, 'porcentaje_avance': 0}

        # 1. Obtener miembros y metas del equipo ECOMMERCE
        equipos_guardados = supabase_manager.read_equipos()        
        ecommerce_vendor_ids = [str(vid) for vid in equipos_guardados.get('ecommerce', [])]
        
        if ecommerce_vendor_ids:
            # 2. Obtener la meta total de ECOMMERCE desde las metas por línea
            meta_ecommerce = metas_del_mes.get('ecommerce', 0)
            kpis_ecommerce['meta_total'] = meta_ecommerce

            # 3. Calcular ventas del equipo ECOMMERCE, agrupadas por LÍNEA COMERCIAL
            ventas_por_linea_ecommerce = {}
            for sale in sales_data:
                user_info = sale.get('invoice_user_id')
                if user_info and isinstance(user_info, list) and len(user_info) > 1:
                    vendedor_id = str(user_info[0])
                    # Si la venta pertenece a un vendedor de ECOMMERCE
                    if vendedor_id in ecommerce_vendor_ids:
                        balance = float(sale.get('balance', 0))
                        
                        # Agrupar por línea comercial con normalización
                        linea_info = sale.get('commercial_line_national_id')
                        linea_nombre = 'N/A'
                        if linea_info and isinstance(linea_info, list) and len(linea_info) > 1:
                            linea_nombre_original = linea_info[1].upper()
                            # Aplicar normalización para agrupar GENVET y MARCA BLANCA como TERCEROS
                            linea_nombre = normalizar_linea_comercial(linea_nombre_original)
                        
                        ventas_por_linea_ecommerce[linea_nombre] = ventas_por_linea_ecommerce.get(linea_nombre, 0) + balance

            # 4. Construir la tabla de datos para la plantilla
            for linea, venta in ventas_por_linea_ecommerce.items():
                datos_ecommerce.append({
                    'nombre': linea, # Ahora es el nombre de la línea comercial
                    'venta': venta
                })
                kpis_ecommerce['venta_total'] += venta

            # 5. Calcular el porcentaje de avance total del equipo
            if kpis_ecommerce['meta_total'] > 0:
                kpis_ecommerce['porcentaje_avance'] = (kpis_ecommerce['venta_total'] / kpis_ecommerce['meta_total']) * 100

            # 6. Calcular el porcentaje de participación de cada línea sobre el total del equipo
            if kpis_ecommerce['venta_total'] > 0:
                for linea_data in datos_ecommerce:
                    linea_data['porcentaje_sobre_total'] = (linea_data['venta'] / kpis_ecommerce['venta_total']) * 100
            else:
                for linea_data in datos_ecommerce:
                    linea_data['porcentaje_sobre_total'] = 0

            # Ordenar las líneas por venta descendente
            datos_ecommerce = sorted(datos_ecommerce, key=lambda x: x['venta'], reverse=True)

        # --- FIN: LÓGICA PARA LA TABLA DEL EQUIPO ECOMMERCE ---

        # Ordenar los datos de la tabla por venta descendente
        datos_lineas_tabla_sorted = sorted(datos_lineas, key=lambda x: x['venta'], reverse=True)
        
        print(f"🔍 DEBUG: datos_lineas tiene {len(datos_lineas)} elementos")
        print(f"🔍 DEBUG: datos_lineas_tabla_sorted tiene {len(datos_lineas_tabla_sorted)} elementos")
        if len(datos_lineas_tabla_sorted) > 0:
            print(f"🔍 DEBUG: Primera línea: {datos_lineas_tabla_sorted[0]}")

        return render_template('dashboard_clean.html',
                             meses_disponibles=meses_disponibles,
                             mes_seleccionado=mes_seleccionado,
                             mes_nombre=mes_nombre,
                             dia_actual=dia_actual,
                             años_disponibles=años_disponibles,
                             año_seleccionado=año_seleccionado,
                             kpis=kpis,
                             datos_lineas=datos_lineas, # Para gráficos, mantener el orden original (alfabético por nombre)
                             datos_lineas_tabla=datos_lineas_tabla_sorted, # Para la tabla, usar los datos ordenados por venta
                             datos_productos=datos_productos,
                             datos_ciclo_vida=datos_ciclo_vida if 'datos_ciclo_vida' in locals() else [],
                             fecha_actual=fecha_actual,
                             avance_lineal_pct=avance_lineal_pct,
                             faltante_meta=faltante_meta,
                             avance_lineal_ipn_pct=avance_lineal_ipn_pct,
                             faltante_meta_ipn=faltante_meta_ipn,
                             datos_ecommerce=datos_ecommerce,
                             kpis_ecommerce=kpis_ecommerce,
                             is_admin=is_admin) # Pasar el flag a la plantilla
    
    except Exception as e:
        # Log completo para debugging (solo en logs, no visible al usuario)
        logger.error(f"Error al obtener datos del dashboard: {type(e).__name__}: {e}", exc_info=True)
        
        # Mensaje genérico para el usuario (sin exponer detalles técnicos)
        flash('No se pudieron cargar los datos del dashboard. Por favor, intente nuevamente.', 'danger')
        
        # Crear datos por defecto para evitar errores
        fecha_actual = datetime.now()
        kpis_default = {
            'meta_total': 0,
            'venta_total': 0,
            'porcentaje_avance': 0,
            'meta_ipn': 0,
            'venta_ipn': 0,
            'porcentaje_avance_ipn': 0,
            'vencimiento_6_meses': 0,
            'avance_diario_total': 0,
            'avance_diario_ipn': 0
        }
        
        return render_template('dashboard_clean.html',
                             meses_disponibles=[{
                                 'key': fecha_actual.strftime('%Y-%m'),
                                 'nombre': f"{fecha_actual.strftime('%B')} {fecha_actual.year}"
                             }],
                             mes_seleccionado=fecha_actual.strftime('%Y-%m'),
                             mes_nombre=f"{fecha_actual.strftime('%B').upper()} {fecha_actual.year}",
                             dia_actual=fecha_actual.day,
                             años_disponibles=list(range(2025, fecha_actual.year + 1)),
                             año_seleccionado=fecha_actual.year,
                             kpis=kpis_default,
                             datos_lineas=[], # Se mantiene vacío en caso de error
                             datos_lineas_tabla=[],
                             datos_productos=[],
                             datos_ciclo_vida=[],
                             fecha_actual=fecha_actual,
                             avance_lineal_pct=0,
                             faltante_meta=0,
                             datos_ecommerce=[],
                             kpis_ecommerce={},
                             is_admin=is_admin) # Pasar el flag a la plantilla


@app.route('/dashboard_linea')
def dashboard_linea():
    if 'username' not in session:
        return redirect(url_for('login'))

    # --- Lógica de Permisos de Administrador ---
    username = session.get('username')
    is_admin = permissions_manager.is_admin(username)

    try:
        # --- 1. OBTENER FILTROS ---
        fecha_actual = datetime.now()
        mes_seleccionado = request.args.get('mes', fecha_actual.strftime('%Y-%m'))
        año_actual = fecha_actual.year
        meses_disponibles = get_meses_del_año(año_actual)

        linea_seleccionada_nombre = request.args.get('linea_nombre', 'PETMEDICA') # Default a PETMEDICA si no se especifica

        # --- NUEVA LÓGICA DE FILTRADO POR DÍA ---
        dia_fin_param = request.args.get('dia_fin')
        año_sel, mes_sel = mes_seleccionado.split('-')

        if dia_fin_param:
            try:
                dia_actual = int(dia_fin_param)
                fecha_fin = f"{año_sel}-{mes_sel}-{str(dia_actual).zfill(2)}"
            except (ValueError, TypeError):
                dia_fin_param = None
        
        if not dia_fin_param:
            if mes_seleccionado == fecha_actual.strftime('%Y-%m'):
                dia_actual = fecha_actual.day
            else:
                ultimo_dia_mes = calendar.monthrange(int(año_sel), int(mes_sel))[1]
                dia_actual = ultimo_dia_mes
            fecha_fin = f"{año_sel}-{mes_sel}-{str(dia_actual).zfill(2)}"
        
        fecha_inicio = f"{año_sel}-{mes_sel}-01"
        # --- FIN DE LA NUEVA LÓGICA ---

        # Mapeo de nombre de línea a ID para cargar metas
        mapeo_nombre_a_id = {
            'PETMEDICA': 'petmedica', 'AGROVET': 'agrovet', 'PET NUTRISCIENCE': 'pet_nutriscience',
            'AVIVET': 'avivet', 'OTROS': 'otros',
            'TERCEROS': 'terceros', 'INTERPET': 'interpet',
        }
        linea_seleccionada_id = mapeo_nombre_a_id.get(linea_seleccionada_nombre.upper(), 'petmedica')

        # --- 2. OBTENER DATOS ---
        # fecha_inicio y fecha_fin se calculan arriba usando la lógica de dia_fin.
        # Asegurar que fecha_inicio siempre esté definida
        año_sel, mes_sel = mes_seleccionado.split('-')
        fecha_inicio = f"{año_sel}-{mes_sel}-01"
        # Si no se definió fecha_fin arriba (por alguna razón), usar el último día del mes
        if 'fecha_fin' not in locals():
            ultimo_dia = calendar.monthrange(int(año_sel), int(mes_sel))[1]
            fecha_fin = f"{año_sel}-{mes_sel}-{ultimo_dia}"

        # Cargar metas de vendedores para el mes y línea seleccionados
        # La estructura es metas[equipo_id][vendedor_id][mes_key]
        metas_vendedores_historicas_raw = supabase_manager.read_metas()
        # Normalizar las claves de equipo_id a minúsculas para que coincidan con linea_seleccionada_id
        metas_vendedores_historicas = {k.lower(): v for k, v in metas_vendedores_historicas_raw.items()}
        # 1. Obtener todas las metas del equipo/línea
        metas_del_equipo = metas_vendedores_historicas.get(linea_seleccionada_id, {})

        # Obtener todos los vendedores de Odoo
        todos_los_vendedores = {str(v['id']): v['name'] for v in data_manager.get_all_sellers()}

        # Obtener ventas del mes
        sales_data = data_manager.get_sales_lines(
            date_from=fecha_inicio,
            date_to=fecha_fin,
            limit=10000
        )

        # --- PRE-FILTRAR VENTAS INTERNACIONALES PARA EFICIENCIA ---
        sales_data_processed = []
        for sale in sales_data:
            # Excluir VENTA INTERNACIONAL (exportaciones) por línea comercial
            linea_comercial = sale.get('commercial_line_national_id')
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                if 'VENTA INTERNACIONAL' in linea_comercial[1].upper():
                    continue
            
            # Excluir VENTA INTERNACIONAL por canal de ventas
            canal_ventas = sale.get('sales_channel_id')
            if canal_ventas and isinstance(canal_ventas, list) and len(canal_ventas) > 1:
                nombre_canal = canal_ventas[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_canal or 'INTERNACIONAL' in nombre_canal:
                    continue
            
            sales_data_processed.append(sale)

        # --- 3. PROCESAR Y AGREGAR DATOS POR VENDEDOR ---
        ventas_por_vendedor = {}
        ventas_ipn_por_vendedor = {}
        ventas_vencimiento_por_vendedor = {}
        ventas_por_producto = {}
        ventas_por_ciclo_vida = {}
        ventas_por_forma = {}
        ajustes_sin_vendedor = 0 # Para notas de crédito sin vendedor
        nombres_vendedores_con_ventas = {} # BUGFIX: Guardar nombres de vendedores con ventas

        for sale in sales_data_processed: # Usar los datos pre-filtrados
            linea_comercial = sale.get('commercial_line_national_id')
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                nombre_linea_original = linea_comercial[1].upper()
                # Aplicar normalización para agrupar GENVET y MARCA BLANCA como TERCEROS
                nombre_linea_actual = normalizar_linea_comercial(nombre_linea_original)

                # Filtrar por la línea comercial seleccionada
                if nombre_linea_actual == linea_seleccionada_nombre.upper():
                    balance = float(sale.get('balance', 0))
                    user_info = sale.get('invoice_user_id')

                    # Si hay un vendedor asignado, se procesa normalmente
                    if user_info and isinstance(user_info, list) and len(user_info) > 1:
                        vendedor_id = str(user_info[0])
                        nombres_vendedores_con_ventas[vendedor_id] = user_info[1] # Guardar el nombre

                        # Agrupar ventas totales
                        ventas_por_vendedor[vendedor_id] = ventas_por_vendedor.get(vendedor_id, 0) + balance

                        # Agrupar ventas IPN
                        if sale.get('product_life_cycle') == 'nuevo':
                            ventas_ipn_por_vendedor[vendedor_id] = ventas_ipn_por_vendedor.get(vendedor_id, 0) + balance
                        
                        # Agrupar ventas por vencimiento < 6 meses
                        ruta = sale.get('route_id')
                        if isinstance(ruta, list) and len(ruta) > 0 and ruta[0] in [18, 19]:
                            ventas_vencimiento_por_vendedor[vendedor_id] = ventas_vencimiento_por_vendedor.get(vendedor_id, 0) + balance
                    
                    # Si NO hay vendedor, se agrupa como un ajuste (ej. Nota de Crédito)
                    else:
                        ajustes_sin_vendedor += balance

                    # Agrupar para gráficos (Top Productos, Ciclo Vida, Forma Farmacéutica)
                    # Esto se hace para todas las transacciones de la línea, con o sin vendedor
                    producto_nombre = sale.get('name', '').strip()
                    if producto_nombre:
                        # Limpiar nombres de ATREVIA eliminando indicadores de tamaño/presentación
                        producto_nombre_limpio = limpiar_nombre_atrevia(producto_nombre)
                        ventas_por_producto[producto_nombre_limpio] = ventas_por_producto.get(producto_nombre_limpio, 0) + balance

                    ciclo_vida = sale.get('product_life_cycle', 'No definido')
                    ventas_por_ciclo_vida[ciclo_vida] = ventas_por_ciclo_vida.get(ciclo_vida, 0) + balance

                    forma_farma = sale.get('pharmaceutical_forms_id')
                    nombre_forma = forma_farma[1] if forma_farma and len(forma_farma) > 1 else 'Instrumental'
                    ventas_por_forma[nombre_forma] = ventas_por_forma.get(nombre_forma, 0) + balance

        # --- 4. CONSTRUIR ESTRUCTURA DE DATOS PARA LA PLANTILLA ---
        datos_vendedores = []
        total_meta = 0
        total_venta = 0
        total_meta_ipn = 0
        total_venta_ipn = 0
        total_vencimiento = 0

        # --- 4.1. UNIFICAR VENDEDORES ---
        # Combinar los vendedores oficiales del equipo con los que tuvieron ventas reales en la línea.
        # Esto asegura que mostremos a todos los miembros del equipo (incluso con 0 ventas)
        # y también a cualquier otra persona que haya vendido en esta línea sin ser miembro oficial.
        equipos_guardados = supabase_manager.read_equipos()
        miembros_oficiales_ids = {str(vid) for vid in equipos_guardados.get(linea_seleccionada_id, [])}
        vendedores_con_ventas_ids = set(ventas_por_vendedor.keys())
        
        todos_los_vendedores_a_mostrar_ids = sorted(list(miembros_oficiales_ids | vendedores_con_ventas_ids))

        # --- 4.2. CONSTRUIR LA TABLA DE VENDEDORES ---
        for vendedor_id in todos_los_vendedores_a_mostrar_ids:
            # BUGFIX: Priorizar el nombre de la venta, luego la lista general, y como último recurso el ID.
            vendedor_nombre = nombres_vendedores_con_ventas.get(vendedor_id, 
                                todos_los_vendedores.get(vendedor_id, f"Vendedor ID {vendedor_id}"))

            
            # Obtener ventas (será 0 si es un miembro oficial sin ventas)
            venta = ventas_por_vendedor.get(vendedor_id, 0)
            venta_ipn = ventas_ipn_por_vendedor.get(vendedor_id, 0)
            vencimiento = ventas_vencimiento_por_vendedor.get(vendedor_id, 0)

            # Asignar meta SOLO si el vendedor es un miembro oficial del equipo
            meta = 0
            meta_ipn = 0
            if vendedor_id in miembros_oficiales_ids:
                meta_guardada = metas_del_equipo.get(vendedor_id, {}).get(mes_seleccionado, {})
                meta = float(meta_guardada.get('meta', 0))
                meta_ipn = float(meta_guardada.get('meta_ipn', 0))

            # Añadir la fila del vendedor a la tabla
            datos_vendedores.append({
                'id': vendedor_id,
                'nombre': vendedor_nombre,
                'meta': meta,
                'venta': venta,
                'porcentaje_avance': (venta / meta * 100) if meta > 0 else 0,
                'meta_ipn': meta_ipn,
                'venta_ipn': venta_ipn,
                'porcentaje_avance_ipn': (venta_ipn / meta_ipn * 100) if meta_ipn > 0 else 0,
                'vencimiento_6_meses': vencimiento
            })

            # Sumar a los totales generales de la línea.
            # La meta solo se suma si fue asignada (es decir, si es miembro oficial).
            # La venta se suma siempre.
            total_meta += meta
            total_venta += venta
            total_meta_ipn += meta_ipn
            total_venta_ipn += venta_ipn
            total_vencimiento += vencimiento

        # --- 4.3. AÑADIR AJUSTES SIN VENDEDOR ---
        if ajustes_sin_vendedor != 0:
            datos_vendedores.append({
                'id': 'ajustes',
                'nombre': 'Ajustes y Notas de Crédito (Sin Vendedor)',
                'meta': 0, 'venta': ajustes_sin_vendedor, 'porcentaje_avance': 0,
                'meta_ipn': 0, 'venta_ipn': 0, 'porcentaje_avance_ipn': 0,
                'vencimiento_6_meses': 0
            })
            # Sumar los ajustes al total de ventas de la línea
            total_venta += ajustes_sin_vendedor

        # Añadir porcentaje sobre el total a cada vendedor
        if total_venta > 0:
            for v in datos_vendedores:
                v['porcentaje_sobre_total'] = (v.get('venta', 0) / total_venta) * 100
        else:
            for v in datos_vendedores:
                v['porcentaje_sobre_total'] = 0

        # --- 4.4. FILTRAR VENDEDORES CON VENTA NEGATIVA ---
        # Si un vendedor solo tiene notas de crédito (venta < 0), no se muestra en la tabla,
        # pero su valor ya fue sumado (restado) al total_venta para mantener la consistencia.
        datos_vendedores_final = [v for v in datos_vendedores if v['venta'] >= 0 or v['id'] == 'ajustes']

        # Ordenar por venta descendente
        datos_vendedores_final = sorted(datos_vendedores_final, key=lambda x: x['venta'], reverse=True)

        # --- 5. CALCULAR KPIs DE LÍNEA ---
        ritmo_diario_requerido_linea = 0
        if mes_seleccionado == fecha_actual.strftime('%Y-%m'):
            hoy = fecha_actual.day
            ultimo_dia_mes = calendar.monthrange(año_actual, fecha_actual.month)[1]
            dias_restantes = 0
            for dia in range(hoy, ultimo_dia_mes + 1):
                if datetime(año_actual, fecha_actual.month, dia).weekday() < 6: # L-S
                    dias_restantes += 1
            
            porcentaje_restante = 100 - ((total_venta / total_meta * 100) if total_meta > 0 else 100)
            if porcentaje_restante > 0 and dias_restantes > 0:
                ritmo_diario_requerido_linea = porcentaje_restante / dias_restantes

        # KPIs generales para la línea
        kpis = {
            'meta_total': total_meta,
            'venta_total': total_venta,
            'porcentaje_avance': (total_venta / total_meta * 100) if total_meta > 0 else 0,
            'meta_ipn': total_meta_ipn,
            'venta_ipn': total_venta_ipn,
            'porcentaje_avance_ipn': (total_venta_ipn / total_meta_ipn * 100) if total_meta_ipn > 0 else 0,
            'vencimiento_6_meses': total_vencimiento,
            'avance_diario_total': ((total_venta / total_meta * 100) / dia_actual) if total_meta > 0 and dia_actual > 0 else 0,
            'avance_diario_ipn': ((total_venta_ipn / total_meta_ipn * 100) / dia_actual) if total_meta_ipn > 0 and dia_actual > 0 else 0,
            'ritmo_diario_requerido': ritmo_diario_requerido_linea
        }

        # --- Avance lineal específico de la línea: proyección de cierre y faltante ---
        try:
            dias_en_mes = calendar.monthrange(int(año_sel), int(mes_sel))[1]
        except Exception:
            dias_en_mes = 30

        if dia_actual > 0:
            proyeccion_mensual_linea = (total_venta / dia_actual) * dias_en_mes
        else:
            proyeccion_mensual_linea = 0

        avance_lineal_pct = (proyeccion_mensual_linea / total_meta * 100) if total_meta > 0 else 0
        faltante_meta = max(total_meta - total_venta, 0)

        # Cálculos específicos para IPN de la línea
        if dia_actual > 0:
            promedio_diario_ipn_linea = total_venta_ipn / dia_actual
            proyeccion_mensual_ipn_linea = promedio_diario_ipn_linea * dias_en_mes
        else:
            proyeccion_mensual_ipn_linea = 0

        avance_lineal_ipn_pct = (proyeccion_mensual_ipn_linea / total_meta_ipn * 100) if total_meta_ipn > 0 else 0
        faltante_meta_ipn = max(total_meta_ipn - total_venta_ipn, 0)

        # Datos para gráficos
        productos_ordenados = sorted(ventas_por_producto.items(), key=lambda x: x[1], reverse=True)[:7]
        datos_productos = [{'nombre': n, 'venta': v} for n, v in productos_ordenados]

        datos_ciclo_vida = [{'ciclo': c, 'venta': v} for c, v in ventas_por_ciclo_vida.items()]
        datos_forma_farmaceutica = [{'forma': f, 'venta': v} for f, v in ventas_por_forma.items()]

        # --- LÓGICA MEJORADA PARA OBTENER LÍNEAS COMERCIALES DISPONIBLES ---
        # Replicar la misma lógica del dashboard principal para consistencia.
        
        # 1. Obtener metas del mes para incluir líneas con metas pero sin ventas.
        metas_historicas = supabase_manager.read_metas_por_linea()
        metas_del_mes = metas_historicas.get(mes_seleccionado, {}).get('metas', {})
        
        # 2. Unificar líneas desde ventas y metas.
        all_lines_dict = {}

        # Desde ventas (aplicando normalización)
        for sale in sales_data_processed: # Usar datos ya filtrados de ventas internacionales
            linea_obj = sale.get('commercial_line_national_id')
            if linea_obj and isinstance(linea_obj, list) and len(linea_obj) > 1:
                linea_nombre_original = linea_obj[1].upper()
                # Aplicar normalización para agrupar GENVET y MARCA BLANCA como TERCEROS
                linea_nombre = normalizar_linea_comercial(linea_nombre_original)
                if linea_nombre not in all_lines_dict:
                    all_lines_dict[linea_nombre] = linea_nombre

        # Desde metas
        for linea_id_meta in metas_del_mes.keys():
            nombre_linea_meta = linea_id_meta.replace('_', ' ').upper()
            if nombre_linea_meta not in all_lines_dict:
                all_lines_dict[nombre_linea_meta] = nombre_linea_meta

        # 3. Filtrar y ordenar
        lineas_a_excluir = ['LICITACION', 'NINGUNO', 'ECOMMERCE', 'VENTA INTERNACIONAL']
        lineas_disponibles = sorted([nombre for nombre in all_lines_dict.values() if nombre not in lineas_a_excluir])
        # --- FIN DE LA LÓGICA MEJORADA ---
        return render_template('dashboard_linea.html',
                               linea_nombre=linea_seleccionada_nombre,
                               mes_seleccionado=mes_seleccionado,
                               meses_disponibles=meses_disponibles,
                               kpis=kpis,
                               datos_vendedores=datos_vendedores_final,
                               datos_productos=datos_productos,
                               datos_ciclo_vida=datos_ciclo_vida,
                               datos_forma_farmaceutica=datos_forma_farmaceutica,
                               lineas_disponibles=lineas_disponibles,
                               fecha_actual=fecha_actual,
                               dia_actual=dia_actual,
                               avance_lineal_pct=avance_lineal_pct,
                               faltante_meta=faltante_meta,
                               avance_lineal_ipn_pct=avance_lineal_ipn_pct,
                               faltante_meta_ipn=faltante_meta_ipn,
                               is_admin=is_admin) # Pasar el flag a la plantilla

    except Exception as e:
        logger.error(f"Error al generar dashboard por línea: {e}", exc_info=True)
        flash('No se pudo cargar el dashboard de la línea comercial. Intente nuevamente.', 'danger')
        # En caso de error, renderizar la plantilla con datos vacíos para no romper la UI
        fecha_actual = datetime.now()
        año_actual = fecha_actual.year
        meses_disponibles = get_meses_del_año(año_actual)
        linea_seleccionada_nombre = request.args.get('linea_nombre', 'PETMEDICA')
        lineas_disponibles = [
            'PETMEDICA', 'AGROVET', 'PET NUTRISCIENCE', 'AVIVET', 'OTROS', 'TERCEROS', 'INTERPET'
        ]
        dia_actual = fecha_actual.day
        kpis_default = {
            'meta_total': 0, 'venta_total': 0, 'porcentaje_avance': 0,
            'meta_ipn': 0, 'venta_ipn': 0, 'porcentaje_avance_ipn': 0,
            'vencimiento_6_meses': 0, 'avance_diario_total': 0, 'avance_diario_ipn': 0
        }
        
        return render_template('dashboard_linea.html',
                               linea_nombre=linea_seleccionada_nombre,
                               mes_seleccionado=fecha_actual.strftime('%Y-%m'),
                               meses_disponibles=meses_disponibles,
                               kpis=kpis_default,
                               datos_vendedores=[],
                               datos_productos=[],
                               datos_ciclo_vida=[],
                               datos_forma_farmaceutica=[],
                               lineas_disponibles=lineas_disponibles,
                               fecha_actual=fecha_actual,
                               dia_actual=dia_actual,
                               avance_lineal_pct=0,
                               faltante_meta=0,
                               avance_lineal_ipn_pct=0,
                               faltante_meta_ipn=0,
                               is_admin=is_admin) # Pasar el flag a la plantilla


@app.route('/meta', methods=['GET', 'POST'])
def meta():
    if 'username' not in session:
        return redirect(url_for('login'))

    # --- Verificación de Permisos ---
    username = session.get('username')
    if not permissions_manager.has_permission(username, 'edit_targets'):
        flash('No tienes permiso para acceder a esta página.', 'warning')
        return redirect(url_for('dashboard'))
    is_admin = permissions_manager.is_admin(username)
    # --- Fin Verificación ---
    
    try:
        # Líneas comerciales estáticas de la empresa
        lineas_comerciales_estaticas = [
            {'nombre': 'PETMEDICA', 'id': 'petmedica'},
            {'nombre': 'AGROVET', 'id': 'agrovet'},
            {'nombre': 'PET NUTRISCIENCE', 'id': 'pet_nutriscience'},
            {'nombre': 'AVIVET', 'id': 'avivet'},
            {'nombre': 'OTROS', 'id': 'otros'},
            {'nombre': 'TERCEROS', 'id': 'terceros'},
            {'nombre': 'INTERPET', 'id': 'interpet'},
        ]
        
        # Obtener año actual y mes seleccionado
        fecha_actual = datetime.now()
        año_actual = fecha_actual.year
        mes_seleccionado = request.args.get('mes', fecha_actual.strftime('%Y-%m'))
        
        # Crear todos los meses del año actual
        meses_año = [{'es_actual': m['key'] == fecha_actual.strftime('%Y-%m'), **m} for m in get_meses_del_año(año_actual)]
        
        if request.method == 'POST':
            # Obtener el mes del formulario
            mes_formulario = request.form.get('mes_seleccionado', mes_seleccionado)
            
            # Procesar metas enviadas
            metas_data = {}
            metas_ipn_data = {}
            total_meta = 0
            total_meta_ipn = 0
            
            for linea in lineas_comerciales_estaticas:
                # Procesar Meta Total
                meta_value = request.form.get(f"meta_{linea['id']}", '0')
                try:
                    # Limpiar formato: remover puntos (separadores de miles) y convertir coma a punto decimal
                    clean_value = str(meta_value).replace('.', '').replace(',', '.') if meta_value else '0'
                    valor = float(clean_value) if clean_value else 0.0
                    metas_data[linea['id']] = valor
                    total_meta += valor
                except (ValueError, TypeError):
                    metas_data[linea['id']] = 0.0
                
                # Procesar Meta IPN
                meta_ipn_value = request.form.get(f"meta_ipn_{linea['id']}", '0')
                try:
                    # Limpiar formato: remover puntos (separadores de miles) y convertir coma a punto decimal
                    clean_value_ipn = str(meta_ipn_value).replace('.', '').replace(',', '.') if meta_ipn_value else '0'
                    valor_ipn = float(clean_value_ipn) if clean_value_ipn else 0.0
                    metas_ipn_data[linea['id']] = valor_ipn
                    total_meta_ipn += valor_ipn
                except (ValueError, TypeError):
                    metas_ipn_data[linea['id']] = 0.0
            
            # --- Procesar Meta ECOMMERCE (campo estático) ---
            # Procesar Meta Total ECOMMERCE
            meta_ecommerce_value = request.form.get('meta_ecommerce', '0')
            try:
                # Limpiar formato: remover puntos (separadores de miles) y convertir coma a punto decimal
                clean_value_ecommerce = str(meta_ecommerce_value).replace('.', '').replace(',', '.') if meta_ecommerce_value else '0'
                valor_ecommerce = float(clean_value_ecommerce) if clean_value_ecommerce else 0.0
                metas_data['ecommerce'] = valor_ecommerce
                total_meta += valor_ecommerce
            except (ValueError, TypeError):
                metas_data['ecommerce'] = 0.0

            # Procesar Meta IPN ECOMMERCE
            meta_ipn_ecommerce_value = request.form.get('meta_ipn_ecommerce', '0')
            try:
                # Limpiar formato: remover puntos (separadores de miles) y convertir coma a punto decimal
                clean_value_ipn_ecommerce = str(meta_ipn_ecommerce_value).replace('.', '').replace(',', '.') if meta_ipn_ecommerce_value else '0'
                valor_ipn_ecommerce = float(clean_value_ipn_ecommerce) if clean_value_ipn_ecommerce else 0.0
                metas_ipn_data['ecommerce'] = valor_ipn_ecommerce
                total_meta_ipn += valor_ipn_ecommerce
            except (ValueError, TypeError):
                metas_ipn_data['ecommerce'] = 0.0
            # --- Fin del procesamiento de ECOMMERCE ---

            # Encontrar el nombre del mes
            mes_obj = next((m for m in meses_año if m['key'] == mes_formulario), None)
            mes_nombre_formulario = mes_obj['nombre'] if mes_obj else ""
            
            # IMPORTANTE: Solo guardar las metas del mes actual, no todo el historial
            metas_solo_mes_actual = {
                mes_formulario: {
                    'metas': metas_data,
                    'metas_ipn': metas_ipn_data,
                    'total': total_meta,
                    'total_ipn': total_meta_ipn,
                    'mes_nombre': mes_nombre_formulario
                }
            }
            supabase_manager.write_metas_por_linea(metas_solo_mes_actual)
            
            flash(f'Metas guardadas exitosamente para {mes_nombre_formulario}. Total: S/ {total_meta:,.0f}', 'success')
            
            # Actualizar mes seleccionado después de guardar
            mes_seleccionado = mes_formulario
        
        # Obtener todas las metas históricas
        metas_historicas = supabase_manager.read_metas_por_linea()
        
        # Obtener metas y total del mes seleccionado
        metas_mes_seleccionado = metas_historicas.get(mes_seleccionado, {})
        metas_actuales_raw = metas_mes_seleccionado.get('metas', {})
        metas_ipn_actuales_raw = metas_mes_seleccionado.get('metas_ipn', {})
        
        # Normalizar claves a minúsculas para coincidir con lineas_comerciales_estaticas
        metas_actuales = {k.lower(): v for k, v in metas_actuales_raw.items()}
        metas_ipn_actuales = {k.lower(): v for k, v in metas_ipn_actuales_raw.items()}
        
        print(f"🔍 DEBUG /meta - Mes seleccionado: {mes_seleccionado}")
        print(f"🔍 DEBUG /meta - Líneas comerciales estáticas: {len(lineas_comerciales_estaticas)}")
        print(f"🔍 DEBUG /meta - Metas actuales ANTES: {metas_actuales}")
        
        # IMPORTANTE: Asegurar que TODAS las líneas comerciales tengan un valor (0 si no existe)
        # Esto permite que se muestren en el formulario aunque no tengan datos previos
        lineas_ids = [linea['id'] for linea in lineas_comerciales_estaticas] + ['ecommerce']
        for linea_id in lineas_ids:
            if linea_id not in metas_actuales:
                metas_actuales[linea_id] = 0.0
            if linea_id not in metas_ipn_actuales:
                metas_ipn_actuales[linea_id] = 0.0
        
        print(f"🔍 DEBUG /meta - Metas actuales DESPUÉS: {metas_actuales}")
        
        total_actual = sum(metas_actuales.values()) if metas_actuales else 0
        total_ipn_actual = sum(metas_ipn_actuales.values()) if metas_ipn_actuales else 0
        
        # Encontrar el nombre del mes seleccionado
        mes_obj_seleccionado = next((m for m in meses_año if m['key'] == mes_seleccionado), meses_año[fecha_actual.month - 1])
        
        return render_template('meta.html',
                             lineas_comerciales=lineas_comerciales_estaticas,
                             metas_actuales=metas_actuales,
                             metas_ipn_actuales=metas_ipn_actuales,
                             metas_historicas=metas_historicas,
                             meses_año=meses_año,
                             mes_seleccionado=mes_seleccionado,
                             mes_nombre=mes_obj_seleccionado['nombre'],
                             total_actual=total_actual,
                             total_ipn_actual=total_ipn_actual,
                             fecha_actual=fecha_actual,
                             is_admin=is_admin) # Pasar el flag a la plantilla
    
    except Exception as e:
        logger.error(f"Error en /meta: {type(e).__name__}: {e}", exc_info=True)
        flash('Error al procesar las metas. Por favor, intente nuevamente.', 'danger')
        return render_template('meta.html',
                             lineas_comerciales=[],
                             metas_actuales={},
                             metas_ipn_actuales={},
                             metas_historicas={},
                             meses_año=[],
                             mes_seleccionado="",
                             mes_nombre="",
                             total_actual=0,
                             total_ipn_actual=0,
                             fecha_actual=datetime.now(),
                             is_admin=is_admin) # Pasar el flag a la plantilla

@app.route('/export/excel/sales')
def export_excel_sales():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    # --- Verificación de Permisos ---
    username = session.get('username')
    if not permissions_manager.has_permission(username, 'export_data'):
        flash('No tienes permiso para exportar datos.', 'warning')
        return redirect(url_for('sales'))
    is_admin = permissions_manager.is_admin(username)
    # --- Fin Verificación ---
    
    try:
        # Obtener filtros de la URL
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        linea_id = request.args.get('linea_id')
        partner_id = request.args.get('partner_id')
        
        # Convertir a tipos apropiados
        if linea_id:
            try:
                linea_id = int(linea_id)
            except (ValueError, TypeError):
                linea_id = None
        
        if partner_id:
            try:
                partner_id = int(partner_id)
            except (ValueError, TypeError):
                partner_id = None
        
        # Obtener datos
        sales_data = data_manager.get_sales_lines(
            date_from=date_from,
            date_to=date_to,
            partner_id=partner_id,
            linea_id=linea_id,
            limit=10000  # Más datos para export
        )
        
        # Filtrar VENTA INTERNACIONAL (exportaciones)
        sales_data_filtered = []
        for sale in sales_data:
            linea_comercial = sale.get('commercial_line_national_id')
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                nombre_linea = linea_comercial[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_linea:
                    continue
            
            # También filtrar por canal de ventas
            canal_ventas = sale.get('sales_channel_id')
            if canal_ventas and isinstance(canal_ventas, list) and len(canal_ventas) > 1:
                nombre_canal = canal_ventas[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_canal or 'INTERNACIONAL' in nombre_canal:
                    continue
            
            sales_data_filtered.append(sale)
        
        # Crear DataFrame
        df = pd.DataFrame(sales_data_filtered)
        
        # Crear archivo Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Ventas', index=False)
        
        output.seek(0)
        
        # Generar nombre de archivo con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'ventas_farmaceuticas_{timestamp}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error al exportar datos: {e}", exc_info=True)
        flash('Error al exportar los datos. Verifique los filtros e intente nuevamente.', 'danger')
        return redirect(url_for('sales'))

@app.route('/metas_vendedor', methods=['GET', 'POST'])
def metas_vendedor():
    if 'username' not in session:
        return redirect(url_for('login'))

    # --- Verificación de Permisos ---
    username = session.get('username')
    if not permissions_manager.has_permission(username, 'edit_targets'):
        flash('No tienes permiso para acceder a esta página.', 'warning')
        return redirect(url_for('dashboard'))
    is_admin = permissions_manager.is_admin(username)
    # --- Fin Verificación ---

    # Obtener meses y líneas comerciales para los filtros
    fecha_actual = datetime.now()
    año_actual = fecha_actual.year
    meses_disponibles = get_meses_del_año(año_actual)
    lineas_comerciales_estaticas = [
        {'nombre': 'PETMEDICA', 'id': 'petmedica'},
        {'nombre': 'AGROVET', 'id': 'agrovet'},
        {'nombre': 'PET NUTRISCIENCE', 'id': 'pet_nutriscience'},
        {'nombre': 'AVIVET', 'id': 'avivet'},
        {'nombre': 'OTROS', 'id': 'otros'},
        {'nombre': 'TERCEROS', 'id': 'terceros'},
        {'nombre': 'INTERPET', 'id': 'interpet'},
    ]
    equipos_definidos = [
        {'id': 'petmedica', 'nombre': 'PETMEDICA'},
        {'id': 'agrovet', 'nombre': 'AGROVET'},
        {'id': 'pet_nutriscience', 'nombre': 'PET NUTRISCIENCE'},
        {'id': 'avivet', 'nombre': 'AVIVET'},
        {'id': 'otros', 'nombre': 'OTROS'},
        {'id': 'terceros', 'nombre': 'TERCEROS'},
        {'id': 'interpet', 'nombre': 'INTERPET'},
    ]

    # Determinar mes y línea seleccionados (desde form o por defecto)
    mes_seleccionado = request.form.get('mes_seleccionado', fecha_actual.strftime('%Y-%m'))
    linea_seleccionada = request.form.get('linea_seleccionada', lineas_comerciales_estaticas[0]['id'])

    if request.method == 'POST':
        # --- 1. GUARDAR ASIGNACIONES DE EQUIPOS ---
        equipo_actualizado_id = request.form.get('guardar_equipo') # Para el mensaje flash
        todos_los_vendedores_para_guardar = data_manager.get_all_sellers()
        equipos_guardados = supabase_manager.read_equipos()

        for equipo in equipos_definidos:
            campo_vendedores = f'vendedores_{equipo["id"]}'
            if campo_vendedores in request.form:
                vendedores_str = request.form.get(campo_vendedores, '')
                if vendedores_str:
                    vendedores_ids = [int(vid) for vid in vendedores_str.split(',') if vid.isdigit()]
                    equipos_guardados[equipo['id']] = vendedores_ids
                else:
                    equipos_guardados[equipo['id']] = []
        supabase_manager.write_equipos(equipos_guardados, todos_los_vendedores_para_guardar)

        # --- 2. GUARDAR SOLO LAS METAS DEL MES SELECCIONADO ---
        # Crear estructura nueva solo con el mes actual, no cargar el histórico completo
        mes_key = mes_seleccionado
        metas_solo_mes_actual = {}
        
        for equipo in equipos_definidos:
            equipo_id = equipo['id']
            metas_solo_mes_actual[equipo_id] = {}

            vendedores_ids_en_equipo = equipos_guardados.get(equipo_id, [])
            for vendedor_id in vendedores_ids_en_equipo:
                vendedor_id_str = str(vendedor_id)

                # Solo procesar el mes actual seleccionado
                meta_valor_str = request.form.get(f'meta_{equipo_id}_{vendedor_id_str}_{mes_key}')
                meta_ipn_valor_str = request.form.get(f'meta_ipn_{equipo_id}_{vendedor_id_str}_{mes_key}')

                # Convertir a float, limpiar formato (puntos y comas)
                try:
                    if meta_valor_str:
                        clean_meta = str(meta_valor_str).replace('.', '').replace(',', '.')
                        meta = float(clean_meta) if clean_meta else None
                    else:
                        meta = None
                except (ValueError, TypeError):
                    meta = None
                
                try:
                    if meta_ipn_valor_str:
                        clean_meta_ipn = str(meta_ipn_valor_str).replace('.', '').replace(',', '.') if meta_ipn_valor_str else '0'
                        meta_ipn = float(clean_meta_ipn) if clean_meta_ipn else None
                    else:
                        meta_ipn = None
                except (ValueError, TypeError):
                    meta_ipn = None

                if meta is not None or meta_ipn is not None:
                    # Solo agregar vendedores que tengan metas asignadas
                    metas_solo_mes_actual[equipo_id][vendedor_id_str] = {
                        mes_key: {
                            'meta': meta or 0.0,
                            'meta_ipn': meta_ipn or 0.0
                        }
                    }

        supabase_manager.write_metas(metas_solo_mes_actual)
        
        if equipo_actualizado_id:
            flash(f'Miembros del equipo actualizados. Ahora puedes asignar sus metas.', 'info')
        else:
            flash('Equipos y metas guardados correctamente.', 'success')

        # Redirigir con los parámetros para recargar la página con los filtros correctos
        return redirect(url_for('metas_vendedor'))

    # GET o después de POST
    todos_los_vendedores = data_manager.get_all_sellers()
    vendedores_por_id = {v['id']: v for v in todos_los_vendedores}
    equipos_guardados = supabase_manager.read_equipos()

    # Construir la estructura de datos para la plantilla
    equipos_con_vendedores = []
    for equipo_def in equipos_definidos:
        equipo_id = equipo_def['id']
        vendedores_ids = equipos_guardados.get(equipo_id, [])
        vendedores_de_equipo = [vendedores_por_id[vid] for vid in vendedores_ids if vid in vendedores_por_id]
        
        equipos_con_vendedores.append({
            'id': equipo_id,
            'nombre': equipo_def['nombre'],
            'vendedores_ids': [str(vid) for vid in vendedores_ids], # Para Tom-Select
            'vendedores': sorted(vendedores_de_equipo, key=lambda v: v['name']) # Para la tabla
        })

    # Para la vista, pasamos todas las metas cargadas
    metas_guardadas = supabase_manager.read_metas()

    return render_template('metas_vendedor.html',
                           meses_disponibles=meses_disponibles,
                           lineas_comerciales=lineas_comerciales_estaticas,
                           equipos_con_vendedores=equipos_con_vendedores,
                           todos_los_vendedores=todos_los_vendedores,
                           metas_guardadas=metas_guardadas,
                           is_admin=is_admin) # Pasar el flag a la plantilla

@app.route('/export/dashboard/details')
def export_dashboard_details():
    """Exporta los detalles del dashboard a un archivo Excel formateado."""
    if 'username' not in session:
        return redirect(url_for('login'))

    # --- Verificación de Permisos ---
    username = session.get('username')
    if not permissions_manager.has_permission(username, 'export_data'):
        flash('No tienes permiso para realizar esta acción.', 'warning')
        return redirect(url_for('dashboard'))
    is_admin = permissions_manager.is_admin(username)
    # --- Fin Verificación ---

    try:
        # Obtener el mes seleccionado de los parámetros de la URL
        mes_seleccionado = request.args.get('mes')
        if not mes_seleccionado:
            flash('No se especificó un mes para la exportación.', 'danger')
            return redirect(url_for('dashboard'))

        # --- Lógica de Fechas (incluyendo filtro de día) ---
        año_sel, mes_sel = mes_seleccionado.split('-')
        fecha_inicio = f"{año_sel}-{mes_sel}-01"

        # Usar el día del parámetro si está disponible, si no, el último día del mes
        dia_fin_param = request.args.get('dia_fin')
        if dia_fin_param and dia_fin_param.isdigit():
            dia_fin = int(dia_fin_param)
            fecha_fin = f"{año_sel}-{mes_sel}-{str(dia_fin).zfill(2)}"
        else:
            # Comportamiento por defecto: mes completo
            ultimo_dia = calendar.monthrange(int(año_sel), int(mes_sel))[1]
            fecha_fin = f"{año_sel}-{mes_sel}-{ultimo_dia}"

        # Obtener datos de ventas reales desde Odoo para ese mes
        sales_data = data_manager.get_sales_lines(
            date_from=fecha_inicio,
            date_to=fecha_fin,
            limit=10000  # Límite alto para exportación
        )

        # Filtrar VENTA INTERNACIONAL (exportaciones), igual que en el dashboard
        sales_data_filtered = []
        for sale in sales_data:
            linea_comercial = sale.get('commercial_line_national_id')
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                if 'VENTA INTERNACIONAL' in linea_comercial[1].upper():
                    continue
            
            canal_ventas = sale.get('sales_channel_id')
            if canal_ventas and isinstance(canal_ventas, list) and len(canal_ventas) > 1:
                nombre_canal = canal_ventas[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_canal or 'INTERNACIONAL' in nombre_canal:
                    continue
            
            sales_data_filtered.append(sale)

        # --- Procesar datos para un formato legible en Excel ---
        processed_for_excel = []
        for record in sales_data_filtered:
            processed_record = {}
            for key, value in record.items():
                # Si el valor es una lista como [id, 'nombre'], extrae solo el nombre
                if isinstance(value, list) and len(value) > 1:
                    processed_record[key] = value[1]
                else:
                    processed_record[key] = value
            
            # Asegurar que el balance sea un número para el formato de moneda
            if 'balance' in processed_record:
                try:
                    processed_record['balance'] = float(processed_record['balance'])
                except (ValueError, TypeError):
                    processed_record['balance'] = 0.0
            
            processed_for_excel.append(processed_record)

        # Crear DataFrame de Pandas con los datos ya procesados
        df = pd.DataFrame(processed_for_excel)

        # --- TRADUCCIÓN Y ORDEN DE COLUMNAS ---
        column_translations = {
            'invoice_date': 'Fecha Factura',
            'l10n_latam_document_type_id': 'Tipo Documento',
            'move_name': 'Número Documento',
            'partner_name': 'Cliente',
            'vat': 'RUC/DNI Cliente',
            'invoice_user_id': 'Vendedor',
            'default_code': 'Código Producto',
            'name': 'Descripción Producto',
            'quantity': 'Cantidad',
            'price_unit': 'Precio Unitario',
            'balance': 'Importe Total',
            'commercial_line_national_id': 'Línea Comercial',
            'sales_channel_id': 'Canal de Venta',
            'payment_state': 'Estado de Pago',
            'invoice_origin': 'Documento Origen',
            'product_life_cycle': 'Ciclo de Vida Producto',
            'pharmacological_classification_id': 'Clasificación Farmacológica',
            'pharmaceutical_forms_id': 'Forma Farmacéutica',
            'administration_way_id': 'Vía de Administración',
            'production_line_id': 'Línea de Producción',
            'categ_id': 'Categoría de Producto',
            'route_id': 'Ruta de Venta'
        }

        # Filtrar el DataFrame para mantener solo las columnas que vamos a usar
        df = df[list(column_translations.keys())]

        # Renombrar las columnas
        df.rename(columns=column_translations, inplace=True)
        
        # El orden de las columnas en el Excel será el mismo que en el diccionario
        # --- FIN DE TRADUCCIÓN Y ORDEN ---

        # --- Creación y Formateo del Archivo Excel ---
        # Crear archivo Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_name = f'Detalle Ventas {mes_seleccionado}'
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Obtener el workbook y la worksheet para aplicar estilos
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # --- Definir Estilos ---
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="875A7B", end_color="875A7B", fill_type="solid")
            currency_format = 'S/ #,##0.00;[Red]-S/ #,##0.00'
            date_format = 'YYYY-MM-DD'
            number_format = '#,##0'

            # --- Aplicar Estilos al Encabezado ---
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill

            # --- Aplicar Formato a Columnas y Ajustar Ancho ---
            for col_idx, column in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                
                # Encontrar el ancho máximo
                if len(df[column]) > 0:
                    max_length = max(df[column].astype(str).map(len).max(), len(column)) + 2
                else:
                    max_length = len(column) + 2
                
                worksheet.column_dimensions[col_letter].width = max_length

                # Aplicar formato a celdas específicas
                if column.lower() == 'balance':
                    for cell in worksheet[col_letter][1:]:
                        cell.number_format = currency_format
                elif column.lower() == 'price_unit':
                    for cell in worksheet[col_letter][1:]:
                        cell.number_format = currency_format
                elif column.lower() == 'quantity':
                    for cell in worksheet[col_letter][1:]:
                        cell.number_format = number_format
                elif 'date' in column.lower():
                    for cell in worksheet[col_letter][1:]:
                        # Convertir texto a objeto datetime si es necesario
                        if isinstance(cell.value, str):
                            try:
                                cell.value = datetime.strptime(cell.value, '%Y-%m-%d')
                            except ValueError:
                                pass # Dejar como texto si no se puede convertir
                        cell.number_format = date_format

            # --- Congelar Panel Superior ---
            worksheet.freeze_panes = 'A2'

        # Mover el cursor al inicio del stream para enviarlo
        output.seek(0)
        # --- Fin del Formateo ---

        # Generar nombre de archivo
        filename = f'detalle_ventas_{mes_seleccionado}.xlsx'

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"Error al exportar detalles del dashboard: {e}", exc_info=True)
        flash('Error al exportar los detalles. Por favor, intente nuevamente.', 'danger')
        return redirect(url_for('dashboard'))


# --- Ruta de Analytics ---
@app.route('/analytics')
def analytics():
    """Dashboard de estadísticas de uso del sistema."""
    if 'username' not in session:
        return redirect(url_for('login'))
    
    # Verificar permisos de administrador
    username = session.get('username')
    if not permissions_manager.has_permission(username, 'view_analytics'):
        flash('No tienes permisos para acceder a esta sección.', 'danger')
        return redirect(url_for('index'))
    
    # Obtener período de análisis
    period = request.args.get('period', '30')
    try:
        days = int(period)
    except ValueError:
        days = 30
    
    # Total de usuarios permitidos (cantidad fija)
    total_allowed_users = 43
    
    # Obtener equipos de ventas desde Supabase
    equipos_guardados = supabase_manager.read_equipos()
    todos_los_vendedores = data_manager.get_all_sellers()
    
    # Crear mapas de email a vendedor_id y vendedor_id a equipo
    email_to_vendedor = {v['name'].lower(): v['id'] for v in todos_los_vendedores if v.get('name')}
    vendedor_to_equipo = {}
    
    # Asignar cada vendedor a su equipo
    for equipo_id, vendedores_ids in equipos_guardados.items():
        for vendedor_id in vendedores_ids:
            # Convertir "OTROS" a "AGROVET" ya que son el mismo equipo
            equipo_nombre = equipo_id.upper()
            if equipo_nombre == 'OTROS':
                equipo_nombre = 'AGROVET'
            vendedor_to_equipo[vendedor_id] = equipo_nombre
    
    # Obtener estadísticas y convertir RealDictRow a diccionarios normales
    stats = {
        'total_visits': analytics_db.get_total_visits(days),
        'unique_users': analytics_db.get_unique_users(days),
        'total_allowed_users': total_allowed_users,
        'visits_by_user': [dict(row) for row in analytics_db.get_visits_by_user(days)],
        'visits_by_page': [dict(row) for row in analytics_db.get_visits_by_page(days)],
        'visits_by_day': [dict(row) for row in analytics_db.get_visits_by_day(days)],
        'visits_by_hour': [dict(row) for row in analytics_db.get_visits_by_hour(min(days, 7))],
        'recent_visits': [dict(row) for row in analytics_db.get_recent_visits(50)]
    }
    
    # Asignar equipo a cada usuario basado en su email
    for visit in stats['visits_by_user']:
        user_email = visit.get('user_email', '')
        user_name = visit.get('user_name', '')
        
        # Intentar encontrar el vendedor por nombre en el email
        vendedor_id = None
        for email_part in [user_email.lower(), user_name.lower()]:
            if email_part in email_to_vendedor:
                vendedor_id = email_to_vendedor[email_part]
                break
        
        # Asignar equipo si se encontró el vendedor
        if vendedor_id and vendedor_id in vendedor_to_equipo:
            visit['equipo'] = vendedor_to_equipo[vendedor_id]
        else:
            visit['equipo'] = ''  # Sin equipo
    
    # Calcular estadísticas por equipo
    equipos_stats = {}
    for visit in stats['visits_by_user']:
        equipo = visit.get('equipo', 'SIN EQUIPO')
        if equipo == '':
            equipo = 'SIN EQUIPO'
        # Convertir "OTROS" a "AGROVET" si aparece
        elif equipo == 'OTROS':
            equipo = 'AGROVET'
        
        if equipo not in equipos_stats:
            equipos_stats[equipo] = {
                'equipo': equipo,
                'total_visitas': 0,
                'usuarios_unicos': 0,
                'usuarios': []
            }
        
        equipos_stats[equipo]['total_visitas'] += visit.get('visit_count', 0)
        equipos_stats[equipo]['usuarios_unicos'] += 1
        equipos_stats[equipo]['usuarios'].append(visit.get('user_name', 'N/A'))
    
    # Convertir a lista y ordenar por visitas
    stats['visits_by_team'] = sorted(
        equipos_stats.values(),
        key=lambda x: x['total_visitas'],
        reverse=True
    )
    
    # Convertir y formatear fechas para compatibilidad con el template
    # IMPORTANTE: Convertir timestamps de UTC a hora de Perú
    for visit in stats['visits_by_user']:
        if visit.get('last_visit'):
            if isinstance(visit['last_visit'], datetime):
                # Asumir que viene en UTC, convertir a Perú
                if visit['last_visit'].tzinfo is None:
                    utc_time = UTC_TZ.localize(visit['last_visit'])
                    peru_time = utc_time.astimezone(PERU_TZ)
                    visit['last_visit'] = peru_time.replace(tzinfo=None)
            elif isinstance(visit['last_visit'], str):
                try:
                    dt = datetime.strptime(visit['last_visit'], '%Y-%m-%d %H:%M:%S')
                    utc_time = UTC_TZ.localize(dt)
                    peru_time = utc_time.astimezone(PERU_TZ)
                    visit['last_visit'] = peru_time.replace(tzinfo=None)
                except:
                    pass
    
    for visit in stats['recent_visits']:
        if visit.get('visit_timestamp'):
            if isinstance(visit['visit_timestamp'], datetime):
                # Asumir que viene en UTC, convertir a Perú
                if visit['visit_timestamp'].tzinfo is None:
                    utc_time = UTC_TZ.localize(visit['visit_timestamp'])
                    peru_time = utc_time.astimezone(PERU_TZ)
                    visit['visit_timestamp'] = peru_time.replace(tzinfo=None)
            elif isinstance(visit['visit_timestamp'], str):
                try:
                    dt = datetime.strptime(visit['visit_timestamp'], '%Y-%m-%d %H:%M:%S')
                    utc_time = UTC_TZ.localize(dt)
                    peru_time = utc_time.astimezone(PERU_TZ)
                    visit['visit_timestamp'] = peru_time.replace(tzinfo=None)
                except:
                    pass
    
    # Ajustar horas de UTC a Perú (restar 5 horas)
    for hour_stat in stats['visits_by_hour']:
        if hour_stat.get('hour') is not None:
            utc_hour = int(hour_stat['hour'])
            peru_hour = (utc_hour - 5) % 24  # Perú es UTC-5
            hour_stat['hour'] = peru_hour
    
    # Preparar datos limpios para los gráficos (solo strings y números)
    chart_labels = []
    chart_visits = []
    chart_unique_users = []
    
    for day in reversed(stats['visits_by_day']):
        if day.get('visit_date'):
            if hasattr(day['visit_date'], 'strftime'):
                chart_labels.append(day['visit_date'].strftime('%d/%m'))
            else:
                chart_labels.append(str(day['visit_date']))
        else:
            chart_labels.append('N/A')
        
        chart_visits.append(day.get('visit_count', 0))
        chart_unique_users.append(day.get('unique_users', 0))
    
    # Generar JavaScript completo en Python
    import json
    chart_js_code = f"""
// Gráfico generado desde Python
if (document.getElementById('visitsPerDayChart')) {{
    new Chart(document.getElementById('visitsPerDayChart'), {{
        type: 'line',
        data: {{
            labels: {json.dumps(chart_labels)},
            datasets: [{{
                label: 'Visitas',
                data: {json.dumps(chart_visits)},
                borderColor: '#875A7B',
                backgroundColor: 'rgba(135, 90, 123, 0.1)',
                tension: 0.4,
                fill: true
            }}, {{
                label: 'Usuarios Únicos',
                data: {json.dumps(chart_unique_users)},
                borderColor: '#00A09D',
                backgroundColor: 'rgba(0, 160, 157, 0.1)',
                tension: 0.4,
                fill: true
            }}]
        }},
        options: {{
            responsive: true,
            maintainAspectRatio: false,
            plugins: {{
                legend: {{
                    position: 'top',
                }}
            }},
            scales: {{
                y: {{
                    beginAtZero: true,
                    ticks: {{
                        precision: 0
                    }}
                }}
            }}
        }}
    }});
}} else {{
    console.error('Canvas visitsPerDayChart no encontrado');
}}
"""
    
    stats['chart_js_code'] = chart_js_code
    
    return render_template('analytics.html', stats=stats, period=days)


# =============================================================================
# MANEJADORES GLOBALES DE ERRORES (ISO 27001 - Sección 5)
# =============================================================================

@app.errorhandler(403)
def forbidden(e):
    """Maneja errores de acceso prohibido (403 Forbidden)"""
    logger.warning(f"Acceso prohibido - Usuario: {session.get('user_email', 'Desconocido')} - Path: {request.path}")
    
    # Si es una petición AJAX, devolver JSON
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'error': 'No tiene permisos para acceder a este recurso'}), 403
    
    # Si es una petición normal, renderizar template
    return render_template('admin/error_403.html'), 403


@app.errorhandler(429)
def ratelimit_handler(e):
    """Maneja errores de rate limiting (429 Too Many Requests)"""
    logger.warning(f"Rate limit excedido - IP: {request.remote_addr} - Path: {request.path} - Usuario: {session.get('username', 'No autenticado')}")
    
    # Si es una petición AJAX, devolver JSON
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({
            'error': 'Demasiados intentos. Por favor, espera un momento.',
            'retry_after': '60 segundos'
        }), 429
    
    # Si es una petición normal, renderizar template
    return render_template('admin/error_429.html'), 429


@app.errorhandler(404)
def page_not_found(e):
    """Maneja errores de página no encontrada (404 Not Found)"""
    logger.info(f"Página no encontrada - Path: {request.path} - Referrer: {request.referrer}")
    
    # Si es una petición AJAX, devolver JSON
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'error': 'Recurso no encontrado'}), 404
    
    # Si es una petición normal, renderizar template
    return render_template('admin/error_404.html'), 404


@app.errorhandler(500)
def internal_error(e):
    """Maneja errores internos del servidor (500 Internal Server Error)"""
    # Log detallado para debugging (con stack trace)
    logger.error(f"Error interno del servidor: {type(e).__name__}: {str(e)}", exc_info=True)
    
    # Si es una petición AJAX, devolver JSON genérico
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # En desarrollo, mostrar detalles; en producción, mensaje genérico
        if app.config.get('DEBUG'):
            return jsonify({'error': f'Error interno: {str(e)}'}), 500
        else:
            return jsonify({'error': 'Error interno del servidor. Por favor, intente nuevamente.'}), 500
    
    # Si es una petición normal, renderizar template
    return render_template('admin/error_500.html'), 500


@app.errorhandler(Exception)
def handle_unexpected_error(e):
    """
    Manejador genérico para errores no manejados.
    Previene exposición de información técnica (ISO 27001 A.14.2.5)
    """
    # Log completo del error con stack trace
    logger.error(f"Error no manejado - Tipo: {type(e).__name__} - Mensaje: {str(e)}", exc_info=True)
    
    # Determinar mensaje de error según contexto
    error_message = 'Ha ocurrido un error. Por favor, intente nuevamente.'
    
    # Mensajes más específicos para ciertos tipos de error
    if 'network' in str(e).lower() or 'connection' in str(e).lower():
        error_message = 'Error de conexión. Verifique su conexión a internet e intente nuevamente.'
    elif 'timeout' in str(e).lower():
        error_message = 'La operación tardó demasiado tiempo. Por favor, intente nuevamente.'
    elif 'permission' in str(e).lower() or 'forbidden' in str(e).lower():
        error_message = 'No tiene permisos para realizar esta operación.'
    
    # Si es una petición AJAX, devolver JSON
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # En desarrollo, incluir tipo de error (sin stack trace)
        if app.config.get('DEBUG'):
            return jsonify({
                'error': error_message,
                'debug_info': f"{type(e).__name__}: {str(e)}"
            }), 500
        else:
            return jsonify({'error': error_message}), 500
    
    # Si es una petición normal, renderizar template
    return render_template('admin/error_500.html', 
                         error_message=error_message if not app.config.get('DEBUG') else str(e)), 500


if __name__ == '__main__':
    logger.info("="*60)
    logger.info("Dashboard de Ventas Farmacéuticas")
    logger.info("Disponible en: http://127.0.0.1:5000")
    logger.info("Usuario: configurado en .env")
    logger.info("="*60)
    app.run(debug=True)
